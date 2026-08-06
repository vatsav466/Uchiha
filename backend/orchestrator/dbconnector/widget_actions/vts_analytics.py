import urdhva_base
import io
import os
import json
import psycopg2
import asyncio
import traceback
import numpy as np
import pandas as pd
import polars as pl
import hpcl_ceg_model
import mysql.connector
from fastapi import Request
import dashboard_studio_model
from datetime import datetime
import polars.selectors as cs
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from hpcl_ceg_model import DeviceInstallation
from fastapi.responses import StreamingResponse
from dateutil.relativedelta import relativedelta
from fastapi.responses import JSONResponse, FileResponse
import pytz
import math
import charts_actions
from orchestrator.dbconnector.widget_actions import widget_actions
import orchestrator.dbconnector.widget_actions.vts_query as vts_query
import orchestrator.dbconnector.credential_loader as credential_loader
import orchestrator.sync_services.vts.vts_ongoing_trips as vts_ongoing_trips
import utilities.connection_mapping as connection_mapping


async def generate_cross_filter(cross_filters):
    _filters, daterange = [], None
    try:
        if cross_filters:
            for f in cross_filters:
                if "DATE" in f.key:
                    start = f.value.split(",")[0]
                    end = f.value.split(",")[-1]
                    daterange = f"'{start} 00:00:00' AND '{end} 23:59:59'"
                else:
                    _filters.append({f.key: f.value})
        return _filters, daterange
    except Exception as e:
        print("--- Exception in cross filters ---")
        print("Exception :", str(e))
        return _filters, daterange

async def get_drill_down_filter(filters, query):
    try:
        conditions = []
        _key = None
        if filters:
            for rec in filters:
                if (rec.key).lower().replace('"', '') in ["rejection_type"]:
                    _key = (rec.value).lower().replace('"', '')
                    continue
                values = rec.value.split(",")
                if len(values) == 1:
                    conditions.append(f'{rec.key} = \'{values[0]}\'')
                else:
                    conditions.append(f"{rec.key} IN {tuple(values)}")
        if conditions:
            if "where" in query.lower():
                query += " AND " + " AND ".join(conditions)
            else:
                query += " WHERE " + " AND ".join(conditions)
        if _key:
            return query, _key
        return query
    except Exception as e:
        print("--- Exception in drill down filters ---")
        print("Exception :", str(e))
        return query

async def filter_data(df, _filters):
    try:        
        if _filters:
            print("-"*30)
            print("_filters :", _filters)
            print("data columns :", df.columns)
            print("length of data :", len(df))
            mask = pd.Series(True, index=df.index)
            for _filter in _filters:
                for key, value in _filter.items():
                    key = key.replace('"','')
                    mask = mask & (df[key].fillna('') == value)
            df = df[mask]
            print("length of filtered data :", len(df))
            print("-"*30)
        return df
    except Exception as e:
        print("Exception in filtering data :", str(e))
    return df

async def safe_json(df):
    return df.replace([np.inf, -np.inf, np.nan], None).to_dict(orient="records")

def extract_invoice_number(invoice_str):
    # Not requried
    """Extract base invoice number: 9017293614-ZF23-1992 -> 9017293614"""
    if pd.isna(invoice_str) or invoice_str is None:
        return None
    invoice_str = str(invoice_str).strip()
    if '-' in invoice_str:
        return invoice_str.split('-')[0]
    return invoice_str


async def get_location_master():
    query = """
        SELECT sap_id, name
        FROM location_master
    """
    df = await VTSAnalyticsActions.execute_query(query, engine="polars")
    return df

async def get_shortage_data(filters, cross_filters, bu_ic, violation_types):
    """Fetch ALL shortage data from sales_trips_till_date"""
    shortage_query = """
                     SELECT vehicle_id, \
                            invoice_no, \
                            plant_nm, \
                            zone_nm, \
                            invoice_date, \
                            destination_code as shortage_destination_code, \
                            COALESCE(NULLIF(TRIM(qty_shortage), 'NaN'), '0.0')::NUMERIC AS qty_shortage, \
                            material_group_nm
                     FROM sales_trips_till_date
                     WHERE load_status in ('6','7') \
                     """

    conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, shortage_query)
    shortage_query = VTSAnalyticsActions.apply_conditions_to_query(shortage_query, conditions)
    df_shortage = await VTSAnalyticsActions.execute_query(shortage_query, engine="polars")

    if df_shortage.is_empty():
        return pl.DataFrame()
    
    # Convert qty_shortage to numeric
    df_shortage = df_shortage.with_columns(
        pl.col('qty_shortage')
        .cast(pl.Float64, strict=False)
        .fill_null(0.0)
        .alias('qty_shortage')
    )

    # Filter out zero shortages for non-I&C BUs when not in default mode
    if not bu_ic and violation_types:
        df_shortage = df_shortage.filter(pl.col("qty_shortage") != 0)
        if df_shortage.is_empty():
            return pl.DataFrame()

    # Create standardized column names for merging
    df_shortage = df_shortage.with_columns(
        invoice_match_key=pl.col("invoice_no").cast(pl.String).str.strip_chars(),
        tl_number=pl.col("vehicle_id"),
        invoice_number=pl.col("invoice_no")
    )

    return df_shortage

async def get_truck_master():
    """Get truck master data for all trucks"""
    truck_query = """
                  SELECT DISTINCT truck_no, \
                                  transporter_name, \
                                  location_name, \
                                  zone
                  FROM vts_truck_master \
                  """
    return await VTSAnalyticsActions.execute_query(truck_query, engine="polars")

def create_shortage_detail(df_shortage: pl.DataFrame, bu_ic: bool, violation_types) -> pl.DataFrame:
    # 1) Clean qty_shortage and material_group_nm
    df_clean = df_shortage.with_columns(
        qty_shortage=(
            pl.col("qty_shortage")
            .cast(pl.Float64, strict=False)  # invalid -> null
            .fill_null(0.0)  # null -> 0
        ),
        material_group_nm=(
            pl.when(pl.col("material_group_nm").is_not_null())
            .then(pl.col("material_group_nm").cast(pl.String).str.strip_chars())
            .otherwise(pl.lit("0"))
        ),
    )

    # 2) Sum shortage per (tl_number, invoice_match_key, material_group_nm)
    per_mat = (
        df_clean
        .group_by(["tl_number", "invoice_match_key", "material_group_nm"])
        .agg(pl.col("qty_shortage").sum().alias("shortage"))
    )

    # 3) Apply "include zeros or only non-zero" rule
    if bu_ic and not violation_types:
        per_mat_filtered = per_mat
    else:
        per_mat_filtered = per_mat.filter(pl.col("shortage") != 0)

    # 4) Build "mat_group:shortage" strings
    per_mat_labeled = per_mat_filtered.with_columns(
        detail=(
                pl.col("material_group_nm").cast(pl.String)
                + pl.lit(":")
                + pl.col("shortage").cast(pl.String)
        )
    )

    # 5) Concatenate details per (tl_number, invoice_match_key)
    df_shortage_grouped = (
        per_mat_labeled
        .group_by(["tl_number", "invoice_match_key"])
        .agg(
            pl.col("detail")
            .str.concat(delimiter=", ")  # join list into a single string
            .alias("qty_shortage_detail")
        )
    )

    return df_shortage_grouped

def determine_merge_strategy(violation_types, bu_ic, bu_tas, bu_lpg):
    """
    Determine merge strategy based on BU and violation_types:
    Returns: (merge_how, base_is_shortage)
    """
    has_violation_filter = bool(violation_types)

    if bu_ic:
        if has_violation_filter:
            return ('inner', True)
        else:
            return ('left', True)
    elif bu_tas or bu_lpg:
        if has_violation_filter:
            return ('left', False)
        else:
            return ('outer', False)
    else:
        return ('left', False)

def merge_shortage_with_violations(
        df_violations, df_shortage, df_truck_master,
        violation_types, bu_ic, bu_tas, bu_lpg,
        all_violations, aggregate=True
):
    """
    Merge shortage and violations based on BU-specific logic
    """
    merge_how, base_is_shortage = determine_merge_strategy(violation_types, bu_ic, bu_tas, bu_lpg)

    if df_shortage.is_empty() and df_violations.is_empty():
        return pl.DataFrame()

    if df_shortage.is_empty():
        if base_is_shortage:
            # If shortage is base and empty, return empty
            return pl.DataFrame()
        else:
            # If violations are base, add empty shortage column
            if aggregate:
                df_violations = df_violations.with_columns(
                    pl.lit(0).cast(pl.Int64).alias("qty_shortage")
                )
            else:
                df_violations = df_violations.with_columns(
                    pl.lit('').alias("qty_shortage_detail")
                )

            # Map with truck master
            if not df_truck_master.is_empty() and not df_violations.is_empty():
                # Prepare lookup: truck_no -> transporter_name
                truck_lookup = (
                    df_truck_master
                    .rename({"truck_no": "tl_number"})  # align join key name
                    .select(["tl_number", "transporter_name"])
                )

                # Ensure transporter_name column exists (like get(..., empty Series))
                if "transporter_name" not in df_violations.columns:
                    df_violations = df_violations.with_columns(
                        transporter_name=pl.lit(None, dtype=pl.String)
                    )

                # Check if we need to fill (any nulls)
                needs_fill = df_violations[
                                 "transporter_name"].null_count() > 0  # cheap metadata op [[Missing data](https://docs.pola.rs/user-guide/expressions/missing-data/#missing-data)]

                if needs_fill:
                    # Join on tl_number to bring in transporter_name from truck master
                    df_violations = (
                        df_violations
                        .join(truck_lookup, on="tl_number", how="left", suffix="_truck")
                        .with_columns(
                            transporter_name=pl.coalesce(
                                "transporter_name",  # existing values
                                "transporter_name_truck",  # from lookup
                            )
                        )
                        .drop("transporter_name_truck")
                    )

            return df_violations

    if df_violations.is_empty():
        if not base_is_shortage:
            # Violations are base but empty
            return pl.DataFrame()

    # Prepare match keys
    if not df_violations.is_empty() and 'invoice_number' in df_violations.columns:
        cleaned_col_expr = pl.col('invoice_number').cast(pl.String).str.strip_chars()
        df_violations = df_violations.with_columns(
            invoice_match_key=pl.when(
                cleaned_col_expr.is_null()
            ).then(
                pl.lit(None)
            ).when(
                cleaned_col_expr.str.contains("-")
            ).then(
                cleaned_col_expr.str.split('-').list.get(0)
            ).otherwise(
                cleaned_col_expr
            )
        )

    df_shortage = df_shortage.with_columns(
        invoice_match_key=pl.col('invoice_no').cast(pl.String).str.strip_chars()
    )

    if aggregate:
        # Aggregate shortage by vehicle + invoice
        df_shortage_agg = (
            df_shortage
            .group_by(["tl_number", "invoice_match_key"])
            .agg(
                pl.col("qty_shortage").sum(),
                pl.col("plant_nm").first(),
                pl.col("zone_nm").first(),
                pl.col("invoice_number").first(),
                pl.col("invoice_date").first(),
                pl.col("shortage_destination_code").first()
            )
        )

        # Perform merge based on strategy
        if base_is_shortage:
            # Shortage is base (I&C)
            if df_violations.is_empty() or 'invoice_number' not in df_violations.columns:
                df_merged = df_shortage_agg.clone()
            else:
                df_merged = df_shortage_agg.join(
                    df_violations,
                    left_on=['tl_number', 'invoice_match_key'],
                    right_on=['tl_number', 'invoice_match_key'],
                    how=merge_how,
                    suffix='_shortage',
                    coalesce=True
                )
                # Use shortage invoice_number when violations invoice_number is null
                if "invoice_number_shortage" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_number=pl.col("invoice_number").fill_null(
                            pl.col("invoice_number_shortage")
                        )
                    )
                else:
                    df_merged = df_merged.with_columns(
                        invoice_number=pl.col("invoice_number").fill_null("")
                    )

                if "invoice_date_shortage" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.col("invoice_date_shortage")
                    )
                elif "invoice_date" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.col("invoice_date")
                    )
                else:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.lit("", dtype=pl.String)
                    )
        else:
            # Violations are base (TAS/LPG/Others)
            if df_violations.is_empty():
                df_merged = df_shortage_agg.clone()
            else:
                df_merged = df_violations.join(
                    df_shortage_agg,
                    left_on=['tl_number', 'invoice_match_key'],
                    right_on=['tl_number', 'invoice_match_key'],
                    how=merge_how,
                    suffix='_shortage',
                    coalesce=True
                )

        # Fill missing data from shortage for unmatched records
        if 'location_name' not in df_merged.columns:
            # Column missing → create using plant_nm or blank
            df_merged = df_merged.with_columns(
                location_name=pl.col("plant_nm").fill_null("") if "plant_nm" in df_merged.columns else pl.lit("")
            )
        else:
            # Column exists → fill NaN with plant_nm
            df_merged = df_merged.with_columns(
                location_name=pl.col("location_name").fill_null(pl.col("plant_nm")).fill_null("")
                if "plant_nm" in df_merged.columns else pl.col("location_name").fill_null("")
            )

        if 'zone' not in df_merged.columns:
            df_merged = df_merged.with_columns(
                zone=pl.col("zone_nm").fill_null("") if "zone_nm" in df_merged.columns else pl.lit("")
            )
        else:
            df_merged = df_merged.with_columns(
                zone=pl.col("zone").fill_null(pl.col("zone_nm")).fill_null("")
                if "zone_nm" in df_merged.columns else pl.col("zone").fill_null("")
            )

        needs_fill = (
                "invoice_number" not in df_merged.columns
                or df_merged["invoice_number"].null_count() > 0
        )

        if needs_fill:
            # ensure invoice_number exists (like get(..., ''))
            if "invoice_number" not in df_merged.columns:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.lit("", dtype=pl.String)
                )

            # if invoice_match_key exists, fill nulls from it; else just fill with ""
            if "invoice_match_key" in df_merged.columns:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.col("invoice_number").fill_null(
                        pl.col("invoice_match_key")
                    )
                )
            else:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.col("invoice_number").fill_null("")
                )

        # Handle created_at/violation_date with fallback to invoice_date
        if 'created_at' not in df_merged.columns:
            if 'violation_date' in df_merged.columns:
                df_merged = df_merged.with_columns(
                    created_at=pl.col("violation_date").fill_null(
                        pl.col("invoice_date")
                    )
                )
            else:
                df_merged = df_merged.with_columns(
                    created_at=pl.col("invoice_date").fill_null("")
                )
        elif "created_at" in df_merged.columns and df_merged["created_at"].null_count() > 0:
            # decide fallback expression: violation_date if present, else invoice_date, else ""
            if "violation_date" in df_merged.columns:
                fallback = pl.col("violation_date")
            elif "invoice_date" in df_merged.columns:
                fallback = pl.col("invoice_date")
            else:
                fallback = pl.lit("", dtype=pl.String)

            df_merged = df_merged.with_columns(
                created_at=pl.col("created_at").fill_null(fallback)
            )

        df_merged = df_merged.with_columns(
            created_at=(
                pl.col("created_at")
                .cast(pl.String)
                .str.to_datetime(format="%Y%m%d", strict=False)
                .dt.strftime("%Y-%m-%d")
            ),
            qty_shortage=(
                pl.col("qty_shortage")
                .cast(pl.Float64, strict=False)
                .fill_null(0)
            ),
        )

    else:
        df_shortage_grouped = create_shortage_detail(
            df_shortage, bu_ic, violation_types
        )

        # Get metadata
        df_shortage_meta = (
            df_shortage
            .group_by(["tl_number", "invoice_match_key"])
            .agg(
                pl.col("plant_nm").first(),
                pl.col("zone_nm").first(),
                pl.col("invoice_number").first(),
                pl.col("invoice_date").first(),
            )
        )

        df_shortage_grouped = df_shortage_grouped.join(
            df_shortage_meta,
            on=['tl_number', 'invoice_match_key'],
            how='left', coalesce=True
        )

        # Perform merge based on strategy
        if base_is_shortage:
            # Shortage is base (I&C)
            if df_violations.is_empty() or 'invoice_number' not in df_violations.columns:
                df_merged = df_shortage_grouped.clone()
            else:
                df_merged = df_shortage_grouped.join(
                    df_violations,
                    left_on=['tl_number', 'invoice_match_key'],
                    right_on=['tl_number', 'invoice_match_key'],
                    how=merge_how,
                    suffix='_shortage',
                    coalesce=True
                )
                if "invoice_number_shortage" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_number=pl.col("invoice_number").fill_null(
                            pl.col("invoice_number_shortage")
                        )
                    )
                else:
                    df_merged = df_merged.with_columns(
                        invoice_number=pl.col("invoice_number").fill_null("")
                    )

                if "invoice_date_shortage" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.col("invoice_date_shortage")
                    )
                elif "invoice_date" in df_merged.columns:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.col("invoice_date")
                    )
                else:
                    df_merged = df_merged.with_columns(
                        invoice_date=pl.lit("", dtype=pl.String)
                    )
        else:
            # Violations are base (TAS/LPG/Others)
            if df_violations.is_empty():
                df_merged = df_shortage_grouped.clone()
            else:
                df_merged = df_violations.join(
                    df_shortage_grouped,
                    left_on=['tl_number', 'invoice_match_key'],
                    right_on=['tl_number', 'invoice_match_key'],
                    how=merge_how,
                    suffix='_shortage',
                    coalesce=True
                )

        # Fill missing data
        if 'location_name' not in df_merged.columns:
            # Column missing → create using plant_nm or blank
            df_merged = df_merged.with_columns(
                location_name=pl.col("plant_nm").fill_null("") if "plant_nm" in df_merged.columns else pl.lit("")
            )
        else:
            df_merged = df_merged.with_columns(
                location_name=pl.col("location_name").fill_null(pl.col("plant_nm"))
                if "plant_nm" in df_merged.columns else pl.col("location_name").fill_null("")
            )

        if 'zone' not in df_merged.columns:
            df_merged = df_merged.with_columns(
                zone=pl.col("zone_nm").fill_null("") if "zone_nm" in df_merged.columns else pl.lit("")
            )
        else:
            df_merged = df_merged.with_columns(
                zone=pl.col("zone").fill_null(pl.col("zone_nm"))
                if "zone_nm" in df_merged.columns else pl.col("zone").fill_null("")
            )

        needs_fill = (
                "invoice_number" not in df_merged.columns
                or df_merged["invoice_number"].null_count() > 0
        )

        if needs_fill:
            # ensure invoice_number exists (like get('invoice_number', ''))
            if "invoice_number" not in df_merged.columns:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.lit("", dtype=pl.String)
                )

            # if invoice_match_key exists, fill nulls from it; else just fill with ""
            if "invoice_match_key" in df_merged.columns:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.col("invoice_number").fill_null(
                        pl.col("invoice_match_key")
                    )
                )
            else:
                df_merged = df_merged.with_columns(
                    invoice_number=pl.col("invoice_number").fill_null("")
                )

        # Handle created_at/violation_date with fallback to invoice_date
        if 'created_at' not in df_merged.columns:
            if 'violation_date' in df_merged.columns:
                df_merged = df_merged.with_columns(
                    created_at=pl.col("violation_date").fill_null(
                        pl.col("invoice_date") if "invoice_date" in df_merged.columns else pl.lit("")
                    )
                )
            else:
                df_merged = df_merged.with_columns(
                    created_at=pl.col("invoice_date") if "invoice_date" in df_merged.columns else pl.lit("")
                )
        elif "created_at" in df_merged.columns and df_merged["created_at"].null_count() > 0:
            # decide fallback: violation_date if present, else invoice_date, else ""
            if "violation_date" in df_merged.columns:
                fallback = pl.col("violation_date")
            elif "invoice_date" in df_merged.columns:
                fallback = pl.col("invoice_date")
            else:
                fallback = pl.lit("", dtype=pl.String)

            df_merged = df_merged.with_columns(
                created_at=pl.col("created_at").fill_null(fallback)
            )

        df_merged = df_merged.with_columns(
            created_at=(
                pl.col("created_at")
                .cast(pl.String)
                .str.to_datetime(format="%Y%m%d", strict=False)
                .dt.strftime("%Y-%m-%d")
            ),
            qty_shortage_detail=(
                pl.col("qty_shortage_detail")
                .cast(pl.String)
                .fill_null("")
            ),
        )

    # Map with truck master for missing transporter_name, location_name, zone
    if not df_truck_master.is_empty() and not df_merged.is_empty():
        # Prepare truck lookup keyed by tl_number
        truck_lookup = (
            df_truck_master
            .rename({"truck_no": "tl_number"})
            .select(["tl_number", "transporter_name", "location_name", "zone"])
        )

        # Ensure transporter_name exists if needed
        if "transporter_name" not in df_merged.columns:
            df_merged = df_merged.with_columns(
                transporter_name=pl.lit(None, dtype=pl.String)
            )

        # Join once to bring in truck info
        df_merged = df_merged.join(
            truck_lookup,
            on="tl_number",
            how="left",
            suffix="_truck",
        )  # gives transporter_name_truck, location_name_truck, zone_truck

        # Fill from truck master where still missing
        df_merged = df_merged.with_columns(
            transporter_name=pl.coalesce(
                "transporter_name", "transporter_name_truck"
            ),
            location_name=pl.coalesce(
                "location_name", "location_name_truck"
            ),
            zone=pl.coalesce(
                "zone", "zone_truck"
            ),
        )

        # Drop helper columns
        df_merged = df_merged.drop(
            ["transporter_name_truck", "location_name_truck", "zone_truck"]
        )

    cols_to_drop = [
        "invoice_match_key",
        "plant_nm",
        "zone_nm",
        "invoice_number_shortage",
        "invoice_date",
        "invoice_date_shortage",
    ]

    existing = [c for c in cols_to_drop if c in df_merged.columns]
    if existing:
        df_merged = df_merged.drop(existing)

    exprs = []
    for v in all_violations:
        if v in df_merged.columns:
            exprs.append(pl.col(v).fill_null(0).alias(v))

    if exprs:
        df_merged = df_merged.with_columns(*exprs)

    return df_merged

async def download_streaming_data(df: pl.DataFrame, filename='violations'):

    df = df.with_columns(
        cs.datetime(time_zone="*").dt.replace_time_zone(None)
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{filename}_{timestamp}.xlsx"

    output = io.BytesIO()
    df.write_excel(
        workbook=output,  # BytesIO
        worksheet=f"{filename}",  # sheet name
    )
    output.seek(0)
    headers = {
        "Content-Disposition": f'attachment; filename="{file_name}"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )

async def streaming_data(df: pl.DataFrame):
    batch_size = 10000
    total = df.height

    async def json_generator():
        for i in range(0, total, batch_size):
            yield json.dumps(df.slice(i, batch_size).to_dicts())
            await asyncio.sleep(0.1)

    return StreamingResponse(
        json_generator(),
        media_type="application/json"
    )

class VTSAnalyticsActions:
    @staticmethod
    def transform_key(key, query=None):
        """Transform keys based on query context"""
        # if query and any(x in query.lower() for x in ["vts_alert_history", "vts_ongoing_trips"]) and key.lower() == "bu":
        #     return "location_type"
        # if query and "vts_alert_history" in query.lower() and key.lower() == "sap_id":
        #     return "location_id" 
        if query and 'sales_trips_till_date' in query.lower() and key.lower() in ('transporter_code','bu'):
            return None
         
        if query and 'sales_trips_till_date' in query.lower() and key.lower() == 'zone':
            return 'zone_nm'
        
        if query and 'sales_trips_till_date' in query.lower() and key.lower() == 'location_name':
            return 'plant_nm'
        
        return key
    
    @staticmethod
    def build_filter_conditions(filters, cross_filters, query):
        """Build WHERE conditions from filters and cross_filters"""
        all_conditions = []
        
        if 'sales_trips_till_date' in query.lower():
            bu_filter = None
            if filters:
                for f in filters:
                    if hasattr(f, 'key') and f.key.lower() == 'bu':
                        bu_filter = f
                        break
            
            bu_value = bu_filter.value if bu_filter else None

            if bu_value:
                bu = bu_value.upper()
                if bu == 'TAS':
                    all_conditions.append("division in ('11', '12','80')")
                    all_conditions.append("sales_org in ('7000','3000','1000')")
                    all_conditions.append("distribution_channel in ('16','12','11')")
                    all_conditions.append("(qty_shortage > '0')")
                elif bu == 'LPG':
                    all_conditions.append("division in ('20', '80')")
                    all_conditions.append("sales_org = '2000'")
                    all_conditions.append("(qty_shortage > '0')")
                elif bu == 'I&C':
                    all_conditions.append("distribution_channel = '12'")
                    all_conditions.append("sales_org = '3000'")
                    all_conditions.append("route <> 'EXW001'")
            
            else:
                all_conditions.append("division = '11'")
                all_conditions.append("sales_org = '7000'")
        
        elif 'sales_trips_till_date' not in query.lower():
            bu_filter = None
            if filters:
                for f in filters:
                    if hasattr(f, 'key') and f.key.lower() == 'bu':
                        bu_filter = f
                        break
            
            if bu_filter and bu_filter.value.upper() == 'I&C':
                all_conditions.append("bu = 'TAS'")
               
        # Process regular filters
        if filters:
            for rec in filters:
                if rec.key.lower() == 'tt_type':
                    all_conditions.append(f"tt_type='{rec.value.lower()}'")
                    continue

                key = VTSAnalyticsActions.transform_key(rec.key, query)
                if key is None:
                    continue

                if (key.lower() == 'bu' and 'sales_trips_till_date' not in query.lower() and
                     rec.value.upper() == 'I&C'):
                   continue

                val = rec.value
                
                condition = VTSAnalyticsActions.create_condition(key, val)
                if condition:
                    all_conditions.append(condition)

        # Process cross filters
        if cross_filters:
            for rec in cross_filters:
                key = rec.key
                val = rec.value

                if "DATE" in key.upper():
                    condition = VTSAnalyticsActions.create_date_condition(query,val)
                else:
                    condition = VTSAnalyticsActions.create_condition(key, val)
                
                if condition:
                    all_conditions.append(condition)
        
        return all_conditions

    @staticmethod
    def create_condition(key, val):
        """Create a single condition based on key and value"""
        if isinstance(val, str):
            if "," in val:  # Handle comma-separated values in string
                values = val.split(",")
                if len(values) == 1:
                    return f"{key} = '{values[0]}'"
                else:
                    return f"{key} IN {tuple(values)}"
            return f"{key} = '{val}'"
        elif isinstance(val, list):
            if len(val) == 1:
                return f"{key} = '{val[0]}'"
            else:
                return f"{key} IN {tuple(val)}"
        return None

    @staticmethod
    def create_date_condition(query,val):
        """Create date range condition"""
        start = val.split(",")[0]
        end_date = val.split(",")[-1]
        end = f"{end_date} 23:59:59"
       
        if "vts_alert_history" in query.lower():
            return f"vts_end_datetime BETWEEN '{start}' AND '{end}'"
        
        if "vts_tripauditmaster" in query.lower():
            return f"createdat BETWEEN '{start}' AND '{end}'"
        
        onging_trips = ["violation_type = 'wr'", "violation_type = 'tc'", "violation_type = 'hs'"]
        if any(ot in query.lower() for ot in onging_trips):
            return f"event_start_datetime BETWEEN '{start}' AND '{end}'"
        
        if "violation_type = 'rd'" in query.lower():
            return f"event_end_datetime BETWEEN '{start}' AND '{end}'"
        
        if "sales_trips_till_date" in query.lower():
            return f"invoice_date::DATE BETWEEN '{start}' AND '{end}'"
        
        if "completed_trips_risk_score" in query.lower():
            return f"insert_datetime BETWEEN '{start}' AND '{end}'"
        
        if "merged_shortage_vts" in query.lower():
            return f"invoice_date BETWEEN '{start}' AND '{end}'"
        
        if any(term in query.lower() for term in ["cluster_master", "transporter_risk_score", "tt_risk_score","clusterwise_event"]):
            return f"version_date BETWEEN '{start}' AND '{end}'"
                
        queries = ["vts_device_removed", "vts_harsh_acceleration", "vts_harsh_braking", "vts_panic"]
        if any(q in query.lower() for q in queries):
            return f"event_date BETWEEN '{start}' AND '{end}'"

        if "non_reporting_devices" in query.lower():
            return f"last_check_date BETWEEN '{start}' AND '{end_date}'"
        
        
        return f"(created_at AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Kolkata')::date BETWEEN '{start}' AND '{end}'"
    

    @staticmethod
    def apply_conditions_to_query(query, conditions):
        """Apply WHERE conditions to query while preserving GROUP BY and ORDER BY"""
        if not conditions:
            return query
        
        conditions_str = " AND ".join(conditions)
        query_lower = query.lower()

        if ") as history_data" in query_lower:
            # Split into two parts: before and after the subquery alias
            idx = query_lower.index(") as history_data")
            subquery_part = query[:idx]  # everything inside the subquery
            rest_part = query[idx:]      # the alias + group by etc.

            # Insert conditions inside the subquery WHERE clause
            if "where" in subquery_part.lower():
                subquery_part = subquery_part.rstrip() + f" AND {conditions_str}"
            else:
                subquery_part = subquery_part.rstrip() + f" WHERE {conditions_str}"

            return subquery_part + rest_part
        
        # Handle queries with GROUP BY
        if "group by" in query_lower:
            idx = query_lower.index("group by")
            base_part = query[:idx].strip()
            group_by_part = query[idx:].strip()
            
            if "where" not in base_part.lower():
                return f"{base_part} WHERE {conditions_str} {group_by_part}"
            else:
                return f"{base_part} AND {conditions_str} {group_by_part}"
        
        # Handle queries with ORDER BY (no GROUP BY)
        elif "order by" in query_lower:
            idx = query_lower.index("order by")
            base_part = query[:idx].strip()
            order_by_part = query[idx:].strip()
            
            if "where" not in base_part.lower():
                return f"{base_part} WHERE {conditions_str} {order_by_part}"
            else:
                return f"{base_part} AND {conditions_str} {order_by_part}"
        
        # Simple query without GROUP BY or ORDER BY
        else:
            if "where" not in query_lower:
                return f"{query} WHERE {conditions_str}"
            else:
                return f"{query} AND {conditions_str}"

    @staticmethod
    def add_alert_type_conditions(conditions, alert_type):
        """Add alert type specific conditions"""
        if not alert_type:
            return conditions
            
        if alert_type.lower() == "blocked":
            conditions.append("alert_status = 'Open' and vehicle_unblocked_date is null") 
        if alert_type.lower() == "acceptance_close":
            conditions.append("alert_status = 'Close' and vehicle_unblocked_date is null")
        elif alert_type.lower() == "auto_unblock":
            conditions.append("alert_status = 'Close' AND mark_as_false = false AND vehicle_unblocked_date is not null")
        elif alert_type.lower() == "manual_unblock":
            conditions.append("alert_status = 'Close' AND mark_as_false = true and vehicle_unblocked_date is not null")
        else:
        # all_alerts or unknown value → no extra condition
           pass
     
        return conditions

    @staticmethod
    def get_period_expression(drill_state):
        """Get period expression based on drill state"""
        if drill_state and drill_state.lower() == "day_wise":
            return "DATE(created_at)"
        elif drill_state and drill_state.lower() == "month_wise":
            return "DATE_TRUNC('month', created_at)"
        return ""

    @staticmethod
    def get_group_by_column(drill_state):
        """Get group by column based on drill state"""
        if drill_state and "location" in drill_state.lower():
            return "location_name"
        elif drill_state and "zone" in drill_state.lower():
            return "zone"
        return None

    @staticmethod
    def format_date(period, drill_state):
        """Format date based on drill state"""
        if drill_state and drill_state.lower() in ["day_wise", "month_wise"]:
            return pd.to_datetime(period).strftime("%b-%d-%Y")
        return str(period)

    @staticmethod
    async def execute_query(query, limit=0, engine='pandas'):
        """Execute query and return DataFrame"""
        try:
            resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=query, limit=limit)
            data = resp.get("data", [])
            if engine == 'polars':
                return pl.DataFrame(data) if data else pl.DataFrame()
            if engine == 'dict':
                return data
            return pd.DataFrame(data) if data else pd.DataFrame()
        except Exception as e:
            print(f"Query execution error: {e}")
            if engine == 'polars':
                return pl.DataFrame()
            if engine == 'dict':
                return []
            return pd.DataFrame()

    @staticmethod
    async def vts_card_chart(filters, cross_filters, drill_state, payload):    
        try:
            # Get base query           
            card_query = vts_query.vts_query.get(drill_state.split(",")[0])

            # Build and apply conditions (pass the query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, card_query)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(card_query, conditions)

            # Execute query
            df = await VTSAnalyticsActions.execute_query(final_query, engine='dict')
            return {"status": True, "message": "success", "data": df}

        except Exception as e:
            print("Exception in BigNumber Chart:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
        
            
    @staticmethod
    async def vts_card_chart_download(filters, cross_filters, drill_state, payload):
        try:
            # Get base query
            card_query = vts_query.vts_query.get(drill_state.split(",")[0])

            # Build conditions
            conditions = VTSAnalyticsActions.build_filter_conditions(
                filters, cross_filters, card_query
            )

            final_query = VTSAnalyticsActions.apply_conditions_to_query(
                card_query, conditions
            )

            # Execute query (assuming returns list of dicts)
            result = await VTSAnalyticsActions.execute_query(final_query)

            # Convert to Polars
            df = pl.DataFrame(result)

            # Extract employee_id and action_by from alert_history index 5
            df = df.with_columns([
                # Safe extraction from alert_history
                pl.col("alert_history")
                .list.get(5, null_on_oob=True)
                .struct.field("employee_id")
                .alias("employee_id"),

                pl.col("alert_history")
                .list.get(5, null_on_oob=True)
                .struct.field("action_by")
                .alias("action_by"),

                # Conditional BU override
                pl.when(pl.col("bu") == "TAS")
                .then(pl.lit("SOD"))
                .otherwise(pl.col("bu"))
                .alias("bu")
            ])

            # Select only required columns (alert_history removed)
            res_df = df.select([
                "bu",
                "sap_id",
                "location_name",
                "unique_id",
                "equipment_name",
                "vehicle_number",
                "created_at",
                "closed_at",
                "employee_id",
                "action_by"
            ])

            return await download_streaming_data(res_df, filename=f'{drill_state}')


        except Exception as e:
            print("Exception in BigNumber Chart:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    
    @staticmethod
    async def vts_dashboard_card_download(filters, cross_filters, drill_state, payload):
        """
            Download VTS dashboard cards as an Excel file.

            This function retrieves the required data from the database and generates
            an Excel file containing the selected columns for download.

            :param filters: Filters applied based on location, SAP ID, and BU for data filtering.
            :param cross_filters: Additional filters, primarily used for date range selection.
            :param drill_state: Used to map and select the appropriate query
                                from the `vts_query` file in the same path.
            :param payload: Additional parameters passed from the frontend
                            for applying extra conditional checks if required.
            :return: Excel file response as a downloadable stream.
            :rtype: dict[str, Any] | StreamingResponse
        """

        try:
            # 1. Build and execute query
            download_card_query = vts_query.vts_query.get(drill_state.split(",")[0])
            conditions = VTSAnalyticsActions.build_filter_conditions(
                filters, cross_filters, download_card_query
            )
            final_query = VTSAnalyticsActions.apply_conditions_to_query(
                download_card_query, conditions
            )

            # df = await VTSAnalyticsActions.execute_query(final_query, engine="polars")
            resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=final_query, limit=0)
            data = resp.get("data", [])
            if not data:
                return {"status": False, "message": "No data found", "data": []}
            
            for row in data:
                for key, value in row.items():
                    if value == "" or value == "[]":
                        row[key] = None


            df = pl.DataFrame(data, infer_schema_length=len(data))
            
            df = df.with_columns([
                   pl.col("alert_history").fill_null([])
             ])

            df = df.with_columns([
                pl.when(pl.col("alert_history").list.len() > 0)
                .then(
                    pl.col("alert_history").list.eval(
                        pl.element().struct.field("employee_id").filter(
                            pl.element().struct.field("action_type") == "Justification"
                        )
                    ).list.first()
                )
                .otherwise(pl.lit(None))
                .alias("creator_id"),

                pl.when(pl.col("alert_history").list.len() > 0)
                .then(
                    pl.col("alert_history").list.eval(
                        pl.element().struct.field("employee_id").filter(
                            pl.element().struct.field("action_type") == "Approved"
                        )
                    ).list.first()
                )
                .otherwise(pl.lit(None))
                .alias("approver_id"),
            ])


            # 3. Select required columns for Excel
            df = (df.select(["zone", "sap_id", "location_name",  "vehicle_number", "violation_type",
                            "unique_id", "alert_status", "device_name", "severity", "created_at", 
                            "creator_id", "approver_id", "vehicle_unblocked_date", "vehicle_blocked_end_date" 
                ])
                .rename({
                    "unique_id": "Alert ID",
                    "sap_id" : "Location ID",
                    "vehicle_number" : "Truck Number",
                    "device_name" : "Instance ID",
                    "creator_id": "Creator ID",
                    "approver_id": "Approver ID",
                })
            )
            # 4. Return Excel download
            return await download_streaming_data(df, filename="itdgAlerts")
        
        except Exception as e:
            print("traceback:", traceback.format_exc())
            print("error",str(e))
            return {"status": False, "message": str(e), "data": []}

    @staticmethod
    async def pagination_df(df, payload):
        total_count = len(df)

        if str(payload.get("download", "")).lower() == "true":
            merged_df = pd.DataFrame(df)
            for col in merged_df.select_dtypes(include=["datetime64[ns, UTC]", "datetimetz"]).columns:
                merged_df[col] = merged_df[col].dt.tz_localize(None)

            merged_df = merged_df.dropna(axis=1, how="all")
            merged_df = merged_df.loc[:, (merged_df.astype(str).apply(lambda x: x.str.strip() != "").any())]

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"violations_{timestamp}.xlsx"

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                merged_df.to_excel(writer, index=False, sheet_name='violations')

            output.seek(0)
            headers = {
                "Content-Disposition": f'attachment; filename="{file_name}"'
            }
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers
            )
        
        skip = int(payload.get("skip", 0))        # page number
        limit = int(payload.get("limit", 20))     # page size
        search_value = str(payload.get("search", "")).strip().lower()
        
        if search_value:
            mask = df.astype(str).apply(
                lambda col: col.str.lower().str.startswith(search_value, na=False)
                ).any(axis=1)
            df = df[mask]

        start = skip * limit
        end = start + limit

        paginated_df = df.iloc[start:end]

        return {
            "status": True,
            "message": "success",
            "total_count": total_count,
            "data": await safe_json(paginated_df)
        }

    @staticmethod
    async def vts_insite(filters, cross_filters, drill_state, payload):
        try:
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            alert_type = payload.get("alert_type") if payload else None
    
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            conditions = VTSAnalyticsActions.add_alert_type_conditions(conditions, alert_type)
            vts_insite_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            print(vts_insite_query)
            df1 = await VTSAnalyticsActions.execute_query(vts_insite_query, engine='polars')

            truck_master_query = """SELECT truck_no, transporter_name FROM vts_truck_master"""
            df_truck_master = await VTSAnalyticsActions.execute_query(truck_master_query, engine='polars')

            merged_df = df1.join(
                    df_truck_master,
                    left_on="vehicle_number",
                    right_on="truck_no",
                    how="left"
                )
            
            merged_df = merged_df.drop(["truck_no"], strict=False)

            if payload.get("download") == "true":
                return await download_streaming_data(merged_df,filename='itdgAlerts')

            return await streaming_data(merged_df)

        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}

    @staticmethod
    async def vts_insite_violation(filters, cross_filters, drill_state, payload):
        """
        Main function to handle VTS violation queries with BU-specific merge logic:
        - I&C: Default=Shortage LEFT Violations, Filtered=Shortage INNER Violations
        - TAS/LPG: Default=Violations OUTER Shortage, Filtered=Violations LEFT Shortage
        - Others: Violations LEFT Shortage (always)
        """
        try:
            # Identify Business Unit
            bu_ic = any(getattr(f, 'key') == 'bu' and getattr(f, 'value') == 'I&C' for f in filters)
            bu_lpg = any(getattr(f, 'key') == 'bu' and getattr(f, 'value') == 'LPG' for f in filters)
            bu_tas = any(getattr(f, 'key') == 'bu' and getattr(f, 'value') == 'TAS' for f in filters)

            query = vts_query.vts_query.get(drill_state.split(",")[0])
            all_violations = vts_query.vts_query.get("all_violations", [])
            violation_types = payload.get("violation_type", []) if payload else []
          

            # Get truck master data (used across all flows)
            df_truck_master = await get_truck_master()

            # Get shortage data and filter for this truck
            df_shortage = await get_shortage_data(filters, cross_filters, bu_ic, violation_types)

            # ==================== DEFAULT CASE - ALL DATA ====================
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            df_history = await VTSAnalyticsActions.execute_query(query, engine="polars")

            if not df_history.is_empty and 'invoice_number' in df_history.columns:
                df_history = df_history.unique(subset=["invoice_number"], keep="first")

            
            # Merge
            final_df = merge_shortage_with_violations(
                df_history, df_shortage, df_truck_master,
                violation_types, bu_ic, bu_tas, bu_lpg,
                all_violations, aggregate=True
            )

            if "shortage_destination_code" in final_df.columns:
                final_df = final_df.with_columns(
                    destination_code=pl.coalesce([
                        pl.col("shortage_destination_code"),  # Prefer shortage
                        pl.col("destination_code")             # Fallback to violation
                    ])
                ).drop("shortage_destination_code")
            
            # loc_df = await get_location_master()
            # final_df = final_df.join(
            #     loc_df.select([
            #         pl.col("sap_id").alias("destination_code"),
            #         pl.col("name").alias("destination_name")
            #     ]),
            #     on="destination_code",
            #     how="left"
            # )
             
            # final_df = final_df.with_columns(
            #   pl.col("destination_name").fill_null("N/A")
            # )
            loc_df = await get_location_master()

            final_df = final_df.join(
                loc_df.select([
                    pl.col("sap_id").alias("destination_code"),
                    pl.col("name").alias("location_master_destination_name")
                ]),
                on="destination_code",
                how="left"
            )
            
            final_df = final_df.with_columns(
                pl.coalesce([
                    pl.col("location_master_destination_name"),
                    pl.col("destination_name")
                ]).alias("destination_name")
            ).drop("location_master_destination_name")
            
            if final_df.is_empty():
                return {"status": True, "message": "No data found", "data": []}
            # Ensure qty_shortage column exists
            if 'qty_shortage' not in final_df.columns:
                final_df = final_df.with_columns(
                    pl.lit(0).cast(pl.Int64).alias("qty_shortage")
                )
            else:
                final_df = final_df.with_columns(
                    qty_shortage=(
                        pl.col("qty_shortage")
                        .cast(pl.Float64, strict=False)
                        .fill_null(0)
                    )
                )

            # Filter out rows with no violations at all
            violation_cols = [v for v in all_violations if v in final_df.columns]
            violation_cols.append('qty_shortage')

            final_df = final_df.with_columns(total_violations=pl.sum_horizontal(pl.col(violation_cols)))

            if bu_ic and not violation_types:
                # For I&C default, include all rows (including zero violations)
                pass
            else:
                final_df = final_df.filter(pl.col("total_violations") > 0)

            if 'total_violations' in final_df.columns:
                final_df = final_df.drop("total_violations")

            if final_df.is_empty():
                return {"status": True, "message": "No violation data found", "data": []}

            # ==================== HANDLE GROUP_BY ====================
            group_by_col = payload.get("group_by") if payload else None
            if group_by_col and group_by_col in final_df.columns:
                violation_cols = [v for v in all_violations if v in final_df.columns]
                # violation_cols.append('qty_shortage')
                agg_df = (
                    final_df
                    .group_by(group_by_col)
                    .agg([pl.col(violation_cols).sum()])
                )

                # add total_count as horizontal sum across violation columns
                agg_df = agg_df.with_columns(
                    total_count=pl.sum_horizontal(pl.col(violation_cols))
                )

                return {
                    "status": True,
                    "message": "success",
                    "data": agg_df.to_dicts()
                }

            # ==================== HANDLE DRILL-DOWN ====================
            qlick_view = payload.get("qlick_view") if payload else None
            click_value = payload.get("click_value") if payload else None
            location_name = payload.get("location_name") if payload else None

            violation_cols = [v for v in all_violations if v in final_df.columns]
            violation_cols.append('qty_shortage')

            # ZONE VIEW
            if qlick_view == "zone" and not click_value:
                final_df = final_df.with_columns(zone=pl.col("zone").fill_null("Unknown"))
                agg_df = (
                    final_df
                    .group_by("zone")
                    .agg([pl.col(violation_cols).sum()])
                )
                agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))
                return {
                    "status": True,
                    "message": "Zone-wise violations",
                    "data": agg_df.to_dicts()
                }

            # ZONE -> LOCATION DRILL
            if qlick_view == "zone" and click_value:
                final_df = final_df.filter(pl.col("zone") == click_value)
                if final_df.is_empty():
                    return {
                        "status": True,
                        "message": f"No data found for zone {click_value}",
                        "data": []
                    }

                agg_df = (
                    final_df
                    .group_by("location_name")
                    .agg([pl.col(violation_cols).sum()])
                )
                agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))

                return {
                    "status": True,
                    "message": f"Violations for all plants in zone {click_value}",
                    "data": agg_df.to_dicts()
                }

            # LOCATION -> TRANSPORTER DRILL
            elif qlick_view == "location_name" and click_value:
                final_df = final_df.filter(pl.col("location_name") == click_value)
                if final_df.is_empty():
                    return {
                        "status": True,
                        "message": f"No data found for location {click_value}",
                        "data": []
                    }
                agg_df = (
                    final_df
                    .group_by("transporter_name")
                    .agg([pl.col(violation_cols).sum()])
                )
                agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))

                return {
                    "status": True,
                    "message": f"Violations for all transporters in location {click_value}",
                    "data": agg_df.to_dicts()
                }

            # TRANSPORTER -> VEHICLE DRILL
            elif qlick_view == "transporter_name" and click_value and location_name:
                final_df = final_df.filter(
                    (pl.col("transporter_name") == click_value) &
                    (pl.col("location_name") == location_name)
                )
                if final_df.is_empty():
                    return {
                        "status": True,
                        "message": f"No data found for transporter {click_value}",
                        "data": []
                    }
                agg_df = (
                    final_df
                    .group_by("tl_number")
                    .agg([pl.col(violation_cols).sum()])
                )
                agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))

                return {
                    "status": True,
                    "message": f"Vehicle-wise violations for transporter {click_value}",
                    "data": agg_df.to_dicts()
                }

            # VEHICLE -> INVOICE DRILL
            elif qlick_view == "tl_number" and click_value:
                final_df = final_df.filter(pl.col("tl_number") == click_value)
                if final_df.is_empty():
                    return {
                        "status": True,
                        "message": f"No data found for vehicle {click_value}",
                        "data": []
                    }

                if 'invoice_number' in final_df.columns:
                    group_cols = ["invoice_number"]
                    if "created_at" in final_df.columns:
                        group_cols.append("created_at")

                    agg_df = (
                        final_df
                        .group_by(group_cols)
                        .agg([pl.col(violation_cols).sum()])
                    )
                    agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))
                else:
                    agg_df = final_df.clone()
                    agg_df = agg_df.with_columns(total_count=pl.sum_horizontal(pl.col(violation_cols)))

                return {
                    "status": True,
                    "message": f"Invoice-wise violations for vehicle {click_value}",
                    "data": agg_df.to_dicts()
                }
            if str(payload.get("download", "")).lower() == "true":
                return await download_streaming_data(final_df, filename='violations')
            return await streaming_data(final_df)

        except Exception as e:
            print("ERROR:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    

    @staticmethod
    async def violation_percentages(filters, cross_filters, drill_state, payload):
        try:
            base_query = vts_query.vts_query.get(drill_state.split(",")[0])

            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query )

            final_query = VTSAnalyticsActions.apply_conditions_to_query(base_query, conditions)

            data = await VTSAnalyticsActions.execute_query(final_query, engine="dict")

            if not data:
                return {"status": True, "message": "No data found", "data": []}

            row = data[0]
            total_trip_count = int(row.get("total_trip_count", 0))

            violation_cols = vts_query.vts_query.get("all_violations", [])

            violation_counts = {
                col: int(row.get(col, 0) or 0)
                for col in violation_cols
            }

        
            emlock_query = vts_query.vts_query.get("emlock_open")
            emlock_conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, emlock_query)
            emlock_query = VTSAnalyticsActions.apply_conditions_to_query(emlock_query, emlock_conditions)
            emlock_data = await VTSAnalyticsActions.execute_query(emlock_query, engine="dict")
            emlock_open = int(emlock_data[0]["emlock_open"]) if emlock_data else 0
            violation_counts["emlock_open"] = emlock_open

        
            shortage_result = await VTSAnalyticsActions.total_count_shortage(filters, cross_filters, drill_state, payload)

            shortage_count = (shortage_result.get("trip_count", 0) if shortage_result.get("status") else 0)
            violation_counts["shortage_count"] = shortage_count

            percentages = {
                key: round(100-(value * 100) / total_trip_count, 2) if total_trip_count > 0 else 0
                for key, value in violation_counts.items()
            }

        
            return {
                "status": True,
                "message": "Violation percentages calculated",
                "data": {
                    "counts": violation_counts,
                    "percentages": percentages,
                    "total_trip": total_trip_count,
                },
            }

        except Exception as e:
            import traceback
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}

    @staticmethod
    async def total_count_shortage(filters, cross_filters, drill_state, payload):
        try:
            # Step 1: Build the base query to get shortage data
            shortage_query = """
            SELECT 
                sap_id, 
                vehicle_id, 
                invoice_no,
                SUM(qty_shortage::numeric) as qty_shortage 
            FROM sales_trips_till_date
            WHERE
                load_status in ('6', '7')
            GROUP BY vehicle_id, invoice_no, sap_id
            """
            
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, shortage_query)
            shortage_query = VTSAnalyticsActions.apply_conditions_to_query(shortage_query, conditions)
            print("Shortage Query:", shortage_query)

            # REMOVED: shortage_query = VTSAnalyticsActions.apply_conditions_to_query(shortage_query, conditions_str)
            
            # Step 2: Get location master data
            location_query = """
            SELECT 
                sap_id,
                bu
            FROM location_master
            """
            
            # Execute both queries
            shortage_df = await VTSAnalyticsActions.execute_query(shortage_query)
            location_df = await VTSAnalyticsActions.execute_query(location_query)
            
            # Check if dataframes are empty
            if shortage_df.empty:
                return {"status": True, "message": "No shortage data found", "data": []}
            
            if location_df.empty:
                return {"status": False, "message": "Location master data not available", "data": []}
            
            # Step 3: Merge the dataframes on sap_id
            merged_df = shortage_df.merge(location_df, on='sap_id', how='left')
            
            filtered_df = merged_df.copy()
            
            if filters:
                for rec in filters:
                    if rec.key.lower() == 'bu':
                        filtered_df = filtered_df[filtered_df['bu'] == rec.value]
            
            total_count = len(filtered_df)
                
            return {
                "status": True, 
                "message": "Success", 
                "trip_count": total_count
            }
            
        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}

    
    @staticmethod
    async def vts_drill_down_violation(filters, cross_filters, drill_state, payload):
        """
        Violation drill-down with pure SQL
        Single query approach - format based on payload, execute, return
        
        Args:
            filters: Base filters (bu, etc.)
            cross_filters: Date range filters
            drill_state: Query key from vts_query
            payload: Contains violation_type and drill parameters
            
        Returns:
            dict with status, message, and data
        """
        try:
            #  Validate violation type
            violation_type = payload.get("violation_type")
            if not violation_type:
                return {"status": False, "message": "violation_type is required", "data": []}
            
            #  Get base query template
            base_query = vts_query.vts_query.get(drill_state.split(",")[0])
            if not base_query:
                return {"status": False, "message": "Invalid drill_state", "data": []}
            
            #  Build SQL parts based on payload
            sql_parts = VTSAnalyticsActions._format_vts_drill_down_query_parts(payload, violation_type, filters, cross_filters)
            
            # Format the query with all parts
            formatted_query = base_query.format(**sql_parts)                    
            #  Execute query
            result_df = await VTSAnalyticsActions.execute_query(formatted_query, engine="polars")
            
            # Step 7: Return results
            if result_df.is_empty():
                return {"status": True, "message": "No data found", "data": []}
            
            return {
                "status": True,
                "message": "Success",
                "data": result_df.to_dicts()
            }
            
        except Exception as e:
            print("Error in vts_drill_down_violation:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}

    
    @staticmethod
    def _format_vts_drill_down_query_parts(payload, violation_type, filters=None, cross_filters=None):
        """
        Build all SQL parts (SELECT, JOIN, FILTERS, GROUP BY, ORDER BY)
        based on payload values
        
        Args:
            payload: Dictionary with drill parameters
            violation_type: Type of violation to analyze
            filters: Optional list of filter objects
            cross_filters: Optional list of cross-filter objects
            
        Returns:
            Dictionary with all formatted SQL parts
        """
        # Extract payload values
        date_wise = str(payload.get("date_wise", "")).lower() == "true"
        tl_number = payload.get("tl_number", "")
        transporter = payload.get("transporter_name", "")
        location = payload.get("location_name", "")
        zone = payload.get("zone", "")
        
        # Check if zone is a boolean flag
        zone_is_true = str(zone).lower() == "true"

        filters = filters or []
        cross_filters = cross_filters or []

        # Build filters from filters list and track what's filtered
        bu_filter = ""
        zone_filter_from_filters = ""
        sap_id_filter = ""
        has_zone_filter = False
        has_sap_id_filter = False
        
        for f in filters:
            if f.key == "bu":
                bu_filter = f"AND vah.bu = '{f.value}'"
            elif f.key == "zone":
                zone_filter_from_filters = f"AND vah.zone = '{f.value}'"
                has_zone_filter = True
            elif f.key == "sap_id":
                sap_id_filter = f"AND vah.sap_id = '{f.value}'"
                has_sap_id_filter = True
        
        # Build date filter from cross_filters
        date_filter = ""
        for cf in cross_filters:
            if cf.key == "DATE":
                date_value = cf.value
                if "," in date_value:
                    start_date, end_date = date_value.split(",")
                    date_filter = f"AND vah.vts_end_datetime BETWEEN '{start_date}' AND '{end_date} 23:59:59'"
                break
        
        # Determine if we need transporter join
        needs_join = bool(transporter or location)
        
        # Build filter clauses for payload-based filters
        # Priority: zone from filters list > zone from payload (if specific value, not "true")
        zone_filter = zone_filter_from_filters if zone_filter_from_filters else (f"AND vah.zone = '{zone}'" if zone and not zone_is_true else "")
        location_filter = f"AND vah.location_name = '{location}'" if location else ""
        transporter_filter = f"AND vtm.transporter_name = '{transporter}'" if transporter else ""
        
        # Mode 1: Invoice Detail (Deepest Drill)
        if tl_number:
            return {
                'violation_type': violation_type,
                'select_clause': f"vah.invoice_number AS invoice_no, DATE(vah.vts_end_datetime) AS created_at, vah.{violation_type}",
                'join_clause': "LEFT JOIN vts_truck_master vtm ON vah.tl_number = vtm.truck_no" if needs_join else "",
                'zone_filter': zone_filter,
                'location_filter': location_filter,
                'transporter_filter': transporter_filter,
                'tl_filter': f"AND vah.tl_number = '{tl_number}'",
                'bu_filter': bu_filter,
                'sap_id_filter': sap_id_filter,
                'date_filter': date_filter,
                'group_clause': f"GROUP BY vah.invoice_number, DATE(vah.vts_end_datetime), vah.{violation_type}",
                'order_clause': "ORDER BY created_at"
            }
        
        # Mode 2 & 3: Aggregation (Summary or Date-wise)
        # Determine drill level and column name based on hierarchy
        if transporter:
            # Drill to vehicle level
            drill_col = "vah.tl_number"
            col_name = "tl_number"
            needs_join = True
        elif location:
            # Drill to transporter level
            drill_col = "vtm.transporter_name"
            col_name = "transporter_name"
            needs_join = True
        elif zone and not zone_is_true:
            # Specific zone value from payload (e.g., "SZ") - drill to location level
            drill_col = "vah.location_name"
            col_name = "location_name"
            needs_join = False
        elif zone_is_true:
            # zone="true" from payload - show zone breakdown
            drill_col = "COALESCE(vah.zone, 'UNKNOWN')"
            col_name = "zone"
            needs_join = False
        elif date_wise:
            # ONLY date_wise, no zone - don't group by zone
            drill_col = None
            col_name = None
            needs_join = False
        else:
            # Default summary - zone level
            drill_col = "COALESCE(vah.zone, 'UNKNOWN')"
            col_name = "zone"
            needs_join = False
        
        # Additional grouping based on filters
        additional_group_cols = []
        additional_select_cols = []
        
        # If zone is in filters, add zone to GROUP BY and SELECT (but avoid duplication with drill_col)
        if has_zone_filter:
            # Only add if drill_col is not already zone-related
            if drill_col not in ["COALESCE(vah.zone, 'UNKNOWN')", "vah.zone"]:
                additional_select_cols.append("vah.zone")
                additional_group_cols.append("vah.zone")
        
        # If sap_id is in filters, add location_name to GROUP BY and SELECT (but avoid duplication)
        if has_sap_id_filter:
            if drill_col != "vah.location_name":
                additional_select_cols.append("vah.location_name")
                additional_group_cols.append("vah.location_name")
        
        # Build SELECT, GROUP BY, ORDER BY dynamically
        if date_wise:
            # Date-wise aggregation
            drill_part = f"{drill_col} AS {col_name}," if drill_col else ""
            additional_select_part = (", ".join(additional_select_cols) + ",") if additional_select_cols else ""
            
            group_drill_part = f", {drill_col}" if drill_col else ""
            additional_group_part = (", " + ", ".join(additional_group_cols)) if additional_group_cols else ""
            
            select_clause = f"""DATE(vah.vts_end_datetime) AS created_at,
                            {additional_select_part}
                            {drill_part} COUNT(DISTINCT vah.invoice_number) AS invoice_count, 
                            COUNT(DISTINCT vah.tl_number) AS vehicle_count, 
                            SUM(vah.{violation_type}) AS {violation_type}"""
            
            group_clause = f"GROUP BY DATE(vah.vts_end_datetime){additional_group_part}{group_drill_part}"
            order_clause = "ORDER BY created_at"
        else:
            # Summary aggregation (no date grouping)
            additional_select_part = (", ".join(additional_select_cols) + ",") if additional_select_cols else ""
            additional_group_part = (", ".join(additional_group_cols)) if additional_group_cols else ""
            
            # Build drill part
            drill_part = f"{drill_col} AS {col_name}," if drill_col else ""
            
            # Combine select parts
            select_clause = f"""{additional_select_part}
                            {drill_part} COUNT(DISTINCT vah.invoice_number) AS invoice_count, 
                            COUNT(DISTINCT vah.tl_number) AS vehicle_count, 
                            SUM(vah.{violation_type}) AS {violation_type}"""
            
            # Build group clause - combine additional and drill columns
            group_parts = []
            if additional_group_part:
                group_parts.append(additional_group_part)
            if drill_col:
                group_parts.append(drill_col)
            
            group_clause = f"GROUP BY {', '.join(group_parts)}" if group_parts else ""
            order_clause = "ORDER BY invoice_count DESC"
        
        return {
            'violation_type': violation_type,
            'select_clause': select_clause,
            'join_clause': "LEFT JOIN vts_truck_master vtm ON vah.tl_number = vtm.truck_no" if needs_join else "",
            'zone_filter': zone_filter,
            'location_filter': location_filter,
            'transporter_filter': transporter_filter,
            'tl_filter': "",
            'bu_filter': bu_filter,
            'sap_id_filter': sap_id_filter,
            'date_filter': date_filter,
            'group_clause': group_clause,
            'order_clause': order_clause
        }


    @staticmethod
    async def vts_drill(filters, cross_filters, drill_state, payload):
        try:
            #  A) DOWNLOAD: MULTIPLE SHEETS BY VIOLATION
            if payload.get("download"):

                violation_query = """
                    SELECT
                        tl_number,
                        invoice_number,
                        location_name,
                        zone,
                        DATE(vts_end_datetime) AS created_at,
                        stoppage_violations_count,
                        route_deviation_count_orig,
                        device_tamper_count,
                        main_supply_removal_count,
                        night_driving_count,
                        speed_violation_count,
                        continuous_driving_count
                    FROM vts_alert_history
                    WHERE 
                        (
                            stoppage_violations_count > 0 OR
                            route_deviation_count_orig > 0 OR
                            device_tamper_count > 0 OR
                            main_supply_removal_count > 0 OR
                            night_driving_count > 0 OR
                            speed_violation_count > 0 OR
                            continuous_driving_count > 0
                        )
                """

                # Apply filters
                conditions = VTSAnalyticsActions.build_filter_conditions(
                    filters, cross_filters, violation_query
                )
                final_query = VTSAnalyticsActions.apply_conditions_to_query(
                    violation_query, conditions
                )

                print("Final violation query for download:", final_query)

                df = await VTSAnalyticsActions.execute_query(final_query)
                df = df.drop_duplicates(keep="first")

                if df is None or df.empty:
                    return {
                        "status": True,
                        "message": "No violation data found for download",
                        "data": []
                    }

                # Violation columns
                violation_cols = [
                    "stoppage_violations_count",
                    "route_deviation_count_orig",
                    "device_tamper_count",
                    "main_supply_removal_count",
                    "night_driving_count",
                    "speed_violation_count",
                    "continuous_driving_count",
                ]

                # Create Excel file
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="xlsxwriter") as writer:

                    # -----------------------------------------------------
                    #  1️⃣  E M L O C K   S H E E T
                    # -----------------------------------------------------
                    print("Generating Emlock Sheet (UI Logic)...")

                    # 1) Fetch UI EMLOCK query
                    emlock_query = vts_query.vts_query.get("get_emlock_open_data")

                    # 2) Apply filters
                    emlock_conditions = VTSAnalyticsActions.build_filter_conditions(
                        filters, cross_filters, emlock_query
                    )
                    emlock_query = VTSAnalyticsActions.apply_conditions_to_query(
                        emlock_query, emlock_conditions
                    )

                    # 3) Execute SQL
                    resp = await urdhva_base.BasePostgresModel.get_aggr_data(
                        query=emlock_query, limit=0
                    )
                    df2 = pd.DataFrame(resp.get("data", []))

                    print("Raw emlock_df count:", len(df2))

                    # 4) Convert TRUE/FALSE to lowercase
                    df2["swipeoutl1"] = df2["swipeoutl1"].fillna("").astype(str).str.lower()
                    df2["swipeoutl2"] = df2["swipeoutl2"].fillna("").astype(str).str.lower()
                    # swipe_out_l1 = df2[df2["swipeoutl1"] == "false"]
                    # swipe_out_l2 = df2[df2["swipeoutl2"] == "false"]

                    final_df_pending = df2[
                        (df2["swipeoutl1"] == "false") |
                        (df2["swipeoutl2"] == "false")
                    ].copy()

                    # final_df_pending = pd.concat([swipe_out_l1, swipe_out_l2]).reset_index(drop=True)
                    final_df_pending["violation_type"] = "open EM Lock"

                    print("Final EMLOCK rows:", len(final_df_pending))

                    # 7) Write to Excel
                    if final_df_pending.empty:
                        pd.DataFrame([{"message": "No data found for emlock_open"}]).to_excel(
                            
                            writer, index=False, sheet_name="emlock_open"
                        )
                    else:
                        final_df_pending.to_excel(writer, index=False, sheet_name="emlock_open")

                    print(f"Wrote {len(final_df_pending)} rows to emlock_open sheet.")
                    

                    # -----------------------------------------------------
                    #  2️⃣  S H O R T A G E   S H E E T
                    # -----------------------------------------------------
                    print("Generating Shortage Sheet...")

                    shortage_query = """
                        SELECT *
                        FROM sales_trips_till_date
                        WHERE load_status in ('6', '7')
                    """
                    shortage_conditions = VTSAnalyticsActions.build_filter_conditions(
                        filters, cross_filters, shortage_query
                    )
                    shortage_query = VTSAnalyticsActions.apply_conditions_to_query(
                        shortage_query, shortage_conditions
                    )

                    shortage_df = await VTSAnalyticsActions.execute_query(shortage_query)

                    if shortage_df is None or shortage_df.empty:
                        pd.DataFrame([{"message": "No data found for shortage_count"}]).to_excel(
                            writer, index=False, sheet_name="shortage_count"
                        )
                    else:
                        shortage_df.to_excel(writer, index=False, sheet_name="shortage_count")

                    # -----------------------------------------------------
                    #  3️⃣  A L L   V I O L A T I O N S   S H E E T
                    # -----------------------------------------------------
                    melt_df = df.melt(
                        id_vars=["tl_number", "invoice_number", "location_name", "zone", "created_at"],
                        value_vars=violation_cols,
                        var_name="violation_type",
                        value_name="violation_value"
                    )

                    # Keep only rows with violations > 0
                    melt_df = melt_df[melt_df["violation_value"] > 0]

                    # Deduplicate: 1 row per invoice per violation TYPE
                    melt_df = melt_df.drop_duplicates(
                        subset=["invoice_number", "violation_type"],
                        keep="first"
                    )

                    # Pivot back to wide format
                    all_violations_df = melt_df.pivot_table(
                        index=["tl_number", "invoice_number", "location_name", "zone", "created_at"],
                        columns="violation_type",
                        values="violation_value",
                        fill_value=0
                    ).reset_index()

                    # Ensure all violation columns are present
                    for v in violation_cols:
                        if v not in all_violations_df.columns:
                            all_violations_df[v] = 0

                    # Sort columns in correct order
                    all_violations_df = all_violations_df[
                        ["tl_number", "invoice_number", "location_name", "zone", "created_at"] + violation_cols
                    ]

                    # Write to Excel
                    all_violations_df.to_excel(writer, index=False, sheet_name="all_violations")

                    # # -----------------------------------------------------
                    # #  4️⃣  S E P A R A T E   V I O L A T I O N   S H E E T S
                    # # -----------------------------------------------------
                    # for col in violation_cols:
                    #     if col not in df.columns:
                    #         continue

                    #     col_df = df[df[col].fillna(0) > 0].copy()
                    #     if col_df.empty:
                    #         continue

                    #     col_df = col_df.sort_values(by="created_at")
                    #     col_df = col_df.drop_duplicates(subset=["invoice_number"], keep="first")

                    #     sheet_name = col[:31]
                    #     col_df.to_excel(writer, index=False, sheet_name=sheet_name)

                # Return file
                output.seek(0)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"vts_violations_{timestamp}.xlsx"

                headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers,
                )

            # ---------------------------------------
            #  B) NORMAL JSON FLOW
            # ---------------------------------------
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            vts_drill_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            print("Drill-down query (non-download):", vts_drill_query)

            vts_df = await VTSAnalyticsActions.execute_query(vts_drill_query)
            vts_df.rename(columns={"vts_end_datetime": "created_at"}, inplace=True)
            vts_df["created_at"] = pd.to_datetime(vts_df["created_at"]).dt.date
            vts_df = vts_df.sort_values(by="created_at", ascending=True)

            if vts_df.empty:
                return {"status": True, "message": "No data found", "data": []}

            transporter_query = """SELECT distinct truck_no, transporter_name FROM vts_truck_master"""
            transporter_df = await VTSAnalyticsActions.execute_query(transporter_query)
            merged_df = vts_df.merge(
                transporter_df, left_on="tl_number", right_on="truck_no", how="left"
            )

            violation_type = payload.get("violation_type")
            if violation_type and violation_type != "all":
                if violation_type not in merged_df.columns:
                    return {
                        "status": False,
                        "message": f"Invalid violation type: {violation_type}",
                        "data": [],
                    }

                violation_filtered_df = merged_df[merged_df[violation_type].fillna(0) != 0].copy()
                violation_filtered_df = violation_filtered_df.sort_values(by="created_at", ascending=True)
                violation_filtered_df = violation_filtered_df.drop_duplicates(
                    subset=["invoice_number"], keep="first"
                )
                data = violation_filtered_df.to_dict(orient="records")
            else:
                data = merged_df.to_dict(orient="records")

            return {"status": True, "message": "success", "data": data}

        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
   
    @staticmethod
    async def non_reporting_devices(filters, cross_filters, drill_state, payload):
        try:
            query = """SELECT zone, location, location_name, truck_regno, loaded_on, (last_check_date || ' ' || last_check_time)::timestamp as last_check,
                       (longitude, latitude) as "longitude/latitude" FROM non_reporting_devices"""
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            
            status_filter = payload.get("status")
            if status_filter == "live":
                conditions.append("completed_trip = 'open' AND completed_trip_auto_dc = 'open'")
            elif status_filter == "closed":
                # conditions.append("'closed' IN (completed_trip, completed_trip_auto_dc)")
                conditions.append("completed_trip='closed' OR completed_trip_auto_dc='closed'")

            query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            df = await VTSAnalyticsActions.execute_query(query)
            if payload.get("total_count") == "true":
                return {"status": True, "total_records":len(df)}
            return {"status": True, "message": "successfull", "data": df.to_dict(orient="records") }
        
        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}


    @staticmethod
    async def vts_ongoing_trips(filters, cross_filters, drill_state, payload):
        try:
            # Step 1: Build and execute base query
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            ongoing_trips_type = payload.get("ongoing_trips_type")

            if ongoing_trips_type:
                query = query.format(ongoing_trips_type=ongoing_trips_type)
            
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)

            status_filter = payload.get("status")
            if status_filter == "live":
                conditions.append("trip_status = 'Live'")
            elif status_filter == "closed":
                conditions.append("trip_status = 'Closed'")

            query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)

            # Step 2: Execute query
            df = await VTSAnalyticsActions.execute_query(query)
            
            for col in ["vehicle_latitude", "vehicle_longitude"]:
                if col in df.columns:
                    df[col] = df[col].astype(str) 

            df = pl.DataFrame(df) if isinstance(df, pd.DataFrame) else df
                      
            if df.is_empty():
                return {"status": True, "message": "No data found", "data": []}
            
            # Remove duplicate records
            df = df.unique(
                subset=["event_start_datetime", "event_end_datetime", "tt_number", "invoice_no"],
                keep="first"
            )

            if df.height == 0:
                return {"status": True, "message": f"No {status_filter} trips found", "data": []}
            
            if payload.get("total_count") == "true":
                return {"status": True, "total_records":df.height}
            
            # Step 4: Handle table view
            if payload.get("table") == "true":
                final_columns = [
                    "event_start_datetime", "event_end_datetime", "sap_id", "region", 
                    "zone", "location_type", "destination_code", "tt_number", "destination_name",
                    "invoice_no", "load_no", "vehicle_latitude", "vehicle_longitude", 
                    "vehicle_location", "transporter_name", "location_name"
                ]
                existing_columns = [col for col in final_columns if col in df.columns]
                table_df = df.select(existing_columns)
                result = table_df.to_dicts()
                return {"status": True, "message": "Data found", "total_records": table_df.height, "data": result}

            # Step 5: TT-level drill-down
            selected_tt = payload.get("tt_number")
            if selected_tt:
                df = df.filter(pl.col("tt_number") == selected_tt)
                
                if df.height == 0:
                    return {"status": True, "message": f"No trips found for vehicle {selected_tt}", "data": []}

                trip_df = df.with_columns(
                    pl.coalesce([
                        pl.col("event_start_datetime"),
                        pl.col("event_end_datetime")
                    ]).cast(pl.Date).cast(str).alias("created_at")
                ).sort("created_at")
                
                trip_df = trip_df.select(["invoice_no", "created_at", "trip_status"])
                result = trip_df.to_dicts()
                
                return {"status": True, "message": f"Trip details for vehicle {selected_tt}", "data": result}

            # Step 6: Fill null values for filters
            df = df.with_columns([
                pl.col("zone").fill_null("Unknown"),
                pl.col("location_name").fill_null("Unknown"),
                pl.col("transporter_name").fill_null("Unknown")
            ])

            # Step 7: Apply payload filters
            if payload.get("zone"):
                df = df.filter(pl.col("zone") == payload["zone"])

            if payload.get("location_name"):
                df = df.filter(pl.col("location_name") == payload["location_name"])

            if payload.get("transporter_name"):
                df = df.filter(pl.col("transporter_name") == payload["transporter_name"])

            # Step 8: Determine grouping column
            if payload.get("transporter_name"):
                group_col = "tt_number"
            elif payload.get("location_name"):
                group_col = "transporter_name"
            elif payload.get("zone"):
                group_col = "location_name"
            else:
                group_col = "zone"

            # Step 9: Create summary
            summary_df = df.group_by(group_col).agg([
                pl.col("invoice_no").n_unique().alias("invoice_count")
            ])

            if group_col != "tt_number":
                vehicle_counts = df.group_by(group_col).agg([
                    pl.col("tt_number").n_unique().alias("vehicle_count")
                ])
                summary_df = summary_df.join(vehicle_counts, on=group_col, how="left")

            # Step 10: Handle Excel download
            if payload.get("download") == "true":
                for col in df.columns:
                    if df[col].dtype in [pl.Datetime, pl.Datetime("ms"), pl.Datetime("us"), pl.Datetime("ns")]:
                        df = df.with_columns(
                            pl.col(col).dt.replace_time_zone(None)
                        )

                violation_mapping = {
                    "HS": "Hotspot",
                    "TC": "Trip not closed more than 2 hours",
                    "RD": "Route Deviation > 2km",
                    "WR": "Trip without route"
                }
                
                if "violation_type" in df.columns:
                    df = df.with_columns(
                        pl.col("violation_type").replace(violation_mapping, default=pl.col("violation_type"))
                    )

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"{ongoing_trips_type}_{status_filter}_{timestamp}.xlsx"
                output = io.BytesIO()
                
                df_pd = df.to_pandas()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df_pd.to_excel(writer, index=False, sheet_name='ongoing_trips')
                
                output.seek(0)
                headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers
                )

            result = summary_df.to_dicts()
            return {"status": True, "message": f"{status_filter} data found", "data": result}

        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}


    @staticmethod
    async def safety_compliance(filters, cross_filters, drill_state, payload):
        try:
            if payload.get("download"):
                print("Download requested — generating Excel with multiple sheets (Polars only)")
                table_names = [tbl.strip() for tbl in drill_state.split(",") if tbl.strip()]
                if not table_names:
                    return JSONResponse({"error": "Missing drill_state (table names)"}, status_code=400)
                output = io.BytesIO()
                for table_name in table_names:
                    query = vts_query.vts_query.get(table_name)
                    query = query.format(drill_state=table_name)

                    conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
                    final_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
                    df = await VTSAnalyticsActions.execute_query(final_query, engine="polars")
                    df = df.unique(keep="first")

                    if df.height == 0:
                        continue

                    sheet_name = table_name[:31]
                    df.write_excel(workbook=output, worksheet=sheet_name)

                output.seek(0)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"safety_complaince_{timestamp}.xlsx"

                headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}

                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers
                )
            query = vts_query.vts_query.get("safety_compliance")
            drill_state_col = drill_state.split(",")[0]
            query = query.format(drill_state=drill_state_col)

            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)

            df = await VTSAnalyticsActions.execute_query(query, engine="polars")
            df = df.unique(keep='first')

            if df.is_empty():
                return {"status": True, "message": "No data found", "data": []}

            for key in ["zone", "location_name", "transporter_name"]:
                if payload.get(key):
                    df = df.filter(pl.col(key) == payload[key])

            if df.is_empty():
                return {"status": True, "message": "No data found for the applied filters", "data": []}
            is_date_wise = str(payload.get("date_wise")).lower() == "true"

            count_col_name = f"{drill_state_col}_count"

            if is_date_wise:
                # detect filter levels
                zone_present = any(f.key == "zone" for f in filters)
                sap_id_present = any(f.key == "sap_id" for f in filters)

                # always start with date
                group_cols = ["created_at"]

                # if zone filter exists → include zone
                if zone_present:
                    group_cols.append("zone")

                # if sap_id exists → include location
                if sap_id_present:
                    group_cols.append("location_name")

                df = (
                    df
                    .with_columns(
                        pl.col("event_date").dt.date().alias("created_at")
                    )
                    .group_by(group_cols)
                    .agg(
                        pl.col("invoice_no").n_unique().alias("invoice_count"),
                        pl.col("tt_number").n_unique().alias("vehicle_count"),
                        pl.col("event_date").count().alias(count_col_name),
                    )
                    .sort("created_at")
                )

                return {
                    "status": True,
                    "message": "Success",
                    "data": df.to_dicts()
                }

            #  TT DETAILS LOGIC
            selected_tt = payload.get("tt_number")
            if selected_tt:
                df = df.filter(pl.col("tt_number") == selected_tt)

                if df.is_empty():
                    return {
                        "status": True,
                        "message": f"No trips found for vehicle {selected_tt}",
                        "data": []
                    }

                df = (
                    df
                    .sort("event_date")
                    .select(["invoice_no", "event_date"])
                    .rename({"event_date": "created_at"})
                )

                return {
                    "status": True,
                    "message": f"Trip details for vehicle {selected_tt}",
                    "data": df.to_dicts()
                }

            #  NORMAL DRILL LOGIC 
            if payload.get("transporter_name"):
                group_col = "tt_number"
            elif payload.get("location_name"):
                group_col = "transporter_name"
            elif payload.get("zone"):
                group_col = "location_name"
            else:
                group_col = "zone"

            df = df.with_columns(pl.col(group_col).fill_null("Unknown"))

            count_col_name = f"{drill_state_col}_count"

            df = (
                df
                .group_by(group_col)
                .agg(
                    pl.col("invoice_no").n_unique().alias("invoice_count"),
                    pl.col("tt_number").n_unique().alias("vehicle_count"),
                    pl.col("event_date").count().alias(count_col_name),
                )
            )

            return {"status": True, "message": "success", "data": df.to_dicts()}

        except Exception as e:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    
                
    @staticmethod
    async def safety_compliance_percentage(filters, cross_filters, drill_state, payload):
                           
            total_trips_count = vts_query.vts_query.get("total_trips")
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, total_trips_count)
            total_trips_count = VTSAnalyticsActions.apply_conditions_to_query(total_trips_count, conditions)            
            df_total = await VTSAnalyticsActions.execute_query(total_trips_count)            
            total_length=int(df_total.iloc[0, 0])
                                      
            try:
                # Queries returning counts
                query_keys = [
                    "vts_panic",
                    "vts_harsh_braking",
                    "vts_harsh_acceleration",
                    "vts_device_removed"
                    
                ]                                                            
                counts = {}
                for key in query_keys:
                    query = vts_query.vts_query.get(key)
                    conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
                    query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
                
                    df = await VTSAnalyticsActions.execute_query(query)
                    counts[key] = int(df.iloc[0, 0]) if not df.empty else 0
                    
                                     
                percentages = {k: round(100-(v / total_length) * 100, 2) for k, v in counts.items()}
                return {"status": True, "message": "Success", "data": { "percentages":
                    percentages,"total_trip":total_length,"counts":counts}}
            

            except Exception as e:
                print("traceback:", traceback.format_exc())
                return {"status": False, "message": str(e), "data": []}
        
            
    @staticmethod
    async def location_level_voilation_breakup(filters, cross_filters, drill_state, payload):
        try:
            # Get query from payload
            query_type = payload.get("query_type") if payload else None
            query = vts_query.vts_query.get(query_type)
            if not query:
                return {"status": False, "message": "Query not found", "data": []}

            # Build and apply conditions (pass the query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)

            # Execute main query
            df = await VTSAnalyticsActions.execute_query(final_query)
            if df.empty:
                return {"status": True, "message": "no data", "data": {}}

            # Fetch location master data
            loc_master_query = "SELECT sap_id, name, zone FROM location_master"
            loc_master_df = await VTSAnalyticsActions.execute_query(loc_master_query)

            # Create location mapping
            loc_map = {}
            for _, row in loc_master_df.iterrows():
                loc_map[row["sap_id"]] = {
                    "zone": row["zone"],
                    "name": row["name"]
                }

            # Build nested data
            nested_data = defaultdict(lambda: defaultdict(int))

            for _, row in df.iterrows():
                location_id = row["location_id"]
                
                # Determine group key
                group_key = None
                if location_id in loc_map:
                    if drill_state and "location" in drill_state.lower():
                        group_key = loc_map[location_id]["name"]
                    elif drill_state and "zone" in drill_state.lower():
                        group_key = loc_map[location_id]["zone"]

                if not group_key:
                    continue

                # Aggregate violation counts
                for col in df.columns:
                    if col != "location_id" and row[col] > 0:
                        nested_data[group_key][col] += row[col]

            # Format final data
            final_data = {}
            for group_key, violations in nested_data.items():
                final_data[group_key] = [
                    {"violation_type": vtype, "count": count}
                    for vtype, count in violations.items()
                ]

            return {"status": True, "message": "success", "data": final_data}

        except Exception as e:
            print("Exception:", str(e))
            return {"status": False, "message": str(e), "data": []}

    @staticmethod
    async def vts_alerts_violations(filters, cross_filters, drill_state, payload):
        try:
            query_type = payload.get("query_type") if payload else None
            alert_type = payload.get("alert_type") if payload else None
            base_query = vts_query.vts_query.get(query_type)
            
            if not base_query:
                return {"status": False, "message": "Query not found", "data": [], "percentages": []}

            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)
            conditions = VTSAnalyticsActions.add_alert_type_conditions(conditions, alert_type)
            
            alerts_query = VTSAnalyticsActions.apply_conditions_to_query(base_query, conditions)

            # Execute queries
            alerts_df = await VTSAnalyticsActions.execute_query(alerts_query, engine='polars')
            
            if alerts_df.is_empty():
                return {"status": True, "message": "success", "data": [], "percentages": []}

            # Get group by column
            group_by_column = VTSAnalyticsActions.get_group_by_column(drill_state)
            if not group_by_column or group_by_column not in alerts_df.columns:
                return {"status": False, "message": f"Column '{group_by_column}' not found", "data": [], "percentages": []}
            
            if 'violation_type' not in alerts_df.columns:
                return {"status": False, "message": "violation_type column not found", "data": [], "percentages": []}
            
            alerts_df = alerts_df.filter((pl.col(group_by_column).is_not_null()) & (pl.col(group_by_column) != ""))
          
            if alerts_df.is_empty():
                return {"status": True, "message": "success", "data": [], "percentages": []}
            
            grouped = (alerts_df.group_by([group_by_column, "violation_type"]).agg(pl.count().alias("count")))
            
            if grouped.is_empty():
                 return {"status": True, "message": "success", "data": [], "percentages": []}
        
            # Prepare response data
            data_response = []
            for group_value in grouped[group_by_column].unique().to_list():
                group_data = grouped.filter(pl.col(group_by_column) == group_value)
                violations_list = [
                    {"violation_type": row['violation_type'], "count": int(row['count'])}
                    for row in group_data.to_dicts()
                ]
                
                if violations_list:
                    data_response.append({group_value: violations_list})

            # Calculate percentages
            violation_totals = (grouped.group_by('violation_type').agg(pl.sum("count")))
            grand_total = violation_totals["count"].sum()
            percentages = []
            if grand_total > 0:
                percentages = [
                    {
                    "violation_type": row["violation_type"],
                    "percentage": round((row["count"] / grand_total) * 100, 2)
                    }
                     for row in violation_totals.iter_rows(named=True)
                ]

            return {
                "status": True, "message": "success",
                "data": data_response, "percentages": percentages
            }

        except Exception as e:
            print("Exception in vts_alerts_violations:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": [], "percentages": []}

    @staticmethod
    async def violation_trends_over_time(filters, cross_filters, drill_state, payload):
        try:
            query_type = payload.get("query_type") if payload else None
            alert_type = payload.get("alert_type") if payload else None
            base_query = vts_query.vts_query.get(query_type)

            if not base_query:
                return {"status": False, "message": "Query not found", "data": []}

            # Build conditions with alert type (pass base_query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)
            conditions = VTSAnalyticsActions.add_alert_type_conditions(conditions, alert_type)

            # Get period expression and format query
            period_expr = VTSAnalyticsActions.get_period_expression(drill_state)
            alerts_query = base_query.format(period_expr=period_expr)
            alerts_query = VTSAnalyticsActions.apply_conditions_to_query(alerts_query, conditions)

            # Execute queries
            alerts_df = await VTSAnalyticsActions.execute_query(alerts_query,engine='polars')
            
            if alerts_df.height == 0:
                return {"status": True, "message": "success", "data": []}

            if not {"violation_type", "period"}.issubset(alerts_df.columns):
                 return {"status": False, "message": "Required columns not found", "data": []}
            
            alerts_df = alerts_df.filter(pl.col("period").is_not_null() & (pl.col("period").cast(pl.Utf8).str.strip_chars() != ""))
            
            if alerts_df.height == 0:
                 return {"status": True, "message": "success", "data": []}
            
            
            grouped = (alerts_df.group_by(['period', 'violation_type']).agg(pl.len().alias("count")))
            
            if grouped.height == 0:
                return {"status": True, "message": "success", "data": []}
            
            result = []
            for period in grouped.select('period').unique().to_series():
                period_data = grouped.filter(pl.col('period') == period)
                formatted_date = VTSAnalyticsActions.format_date(period, drill_state)
                
                values = [
                    {"violation_type": row['violation_type'], "count": int(row['count'])}
                    for row in period_data.iter_rows(named=True)
                ]
                
                if values:
                    result.append({"date": formatted_date, "records": values})
            
            # Sort by period (assuming period is sortable)
            result.sort(key=lambda x: x['date'])
            
            return {"status": True, "message": "success", "data": result}

        except Exception as e:
            print("Exception in violation_trends_over_time:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
        
    @staticmethod
    async def violation_details(filters, cross_filters, drill_state, payload):
        try:
            query_type = payload.get("query_type") if payload else None
            violation_type = payload.get("violation_type")
            base_query = vts_query.vts_query.get(query_type)
            
            if not base_query:
                return {"status": False, "message": "Query not found", "data": []}

            # Build conditions and format query (pass base_query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)
            period_expr = VTSAnalyticsActions.get_period_expression(drill_state)
            
            query = base_query.format(period_expr=period_expr, violation_type=violation_type)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)

            # Execute query
            df = await VTSAnalyticsActions.execute_query(final_query)
            if df.empty:
                return {"status": True, "message": "success", "data": []}

            # Process numeric columns
            numeric_cols = [col for col in df.columns if col != "period"]
            for col in numeric_cols:
                df[col] = df[col].astype(int)
            
            # Split columns into summary and instance
            summary_cols = numeric_cols[:4] 
            instance_cols = numeric_cols[4:]
            
            # Calculate summaries
            summary_counts = [{col: int(df[col].sum()) for col in summary_cols}]
            
            overall_instance_totals = {col: int(df[col].sum()) for col in instance_cols}
            grand_total = sum(overall_instance_totals.values())

            instance_breakup = {}
            for col, total_count in overall_instance_totals.items():
                instance_breakup[col] = {
                    "total_count": total_count,
                    "percentage": round((total_count / grand_total) * 100, 2) if grand_total > 0 else 0
                }

            # Format period data
            period_data = []
            for _, row in df.iterrows():
                counts = {col: int(row[col]) for col in instance_cols}
                formatted_date = VTSAnalyticsActions.format_date(row["period"], drill_state)
                
                period_data.append({
                    "date": formatted_date,
                    "value": {"counts": counts}
                })
        
            return {
                "status": True, "message": "success",
                "data": {
                    violation_type: summary_counts,
                    "period_wise": period_data,
                    "instance_breakup": instance_breakup
                }
            }
        
        except Exception as e:
            print("Exception:", str(e))
            print(traceback.format_exc())
            return {"status": False, "message": str(e), "data": {}}
                 

    @staticmethod
    async def alert_summary(filters, cross_filters, drill_state, payload):
        try:
            query_type = payload.get("query_type") if payload else None
            violation_type = payload.get("violation_type")
            query = vts_query.vts_query.get(query_type)
            
            if not query:
                return {"status": False, "message": "Query not found", "data": []}
            
            # Get group by column and build conditions (pass query for key transformation)
            group_by_column = VTSAnalyticsActions.get_group_by_column(drill_state)
            if not group_by_column:
                return {"status": False, "message": "Invalid drill state", "data": []}
                
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            
            # Format and execute query
            formatted_query = query.format(group_by_column=group_by_column, violation_type=violation_type)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(formatted_query, conditions)

            df = await VTSAnalyticsActions.execute_query(final_query)
            
            # Filter out null/empty values
            df = df[df[group_by_column].notna()]
            df = df[df[group_by_column].astype(str).str.strip() != ""]

            if df.empty:
                return {"status": True, "message": "no data", "data": {}}
           
            # Format results
            final_result = {}
            for _, row in df.iterrows():
                group_val = row[group_by_column] 
                instance = row["instance_level"]

                if group_val not in final_result:
                    final_result[group_val] = []

                final_result[group_val].append({
                    instance: [{
                        "Blocked": row["Blocked"],
                        "Auto Unblock": row["Auto Unblock"],
                        "Manual Unblock": row["Manual Unblock"],
                        "Total": row["Total"]
                    }]
                })

            return {"status": True, "message": "success", "data": final_result}

        except Exception as e:
            return {"status": False, "message": str(e), "data": {}}

    @staticmethod
    async def card_chart_shortage(filters, cross_filters, drill_state, payload):
        try:
            card_query = vts_query.vts_query.get(drill_state.split(",")[0])
            query = vts_query.vts_query.get(card_query)
            
            if not query:
                return {"status": False, "message": "Query not found", "data": []}

            # Build and apply conditions (pass query for key transformation)
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            
            # Execute VTS history query
            vts_history = await VTSAnalyticsActions.execute_query(final_query)

            # Execute Tibco query
            conn = await VTSAnalyticsActions.tibco_connection()
            if not conn:
                return {"status": False, "message": "Database connection failed", "data": {}}
                
            shortage_tibco_query = vts_query.vts_query.get("shortage_tibco")
            shortage_resp = await VTSAnalyticsActions.execute_tibco_query(conn, shortage_tibco_query)
            shortage_data = pd.DataFrame(shortage_resp.get("data", []))
            
            # Process and merge data
            vts_history = vts_history[
                pd.notnull(vts_history["invoice_number"]) & 
                (vts_history["invoice_number"] != "")
            ]
            vts_history["invoice_prefix"] = vts_history["invoice_number"].apply(lambda x: x.split("-")[0])

            merged_df = pd.merge(
                vts_history, shortage_data,
                left_on="invoice_prefix", right_on="INVOICE_NO",
                how="inner"
            )
            
            total_qty_shortage = int(merged_df["QTY_SHORTAGE"].sum())
            
            # Calculate total violations
            violation_cols = [col for col in vts_history.columns if col not in ["invoice_number", "invoice_prefix"]]
            total_violation_count = vts_history[violation_cols].sum().sum()

            shortage_percentage = round((total_qty_shortage / total_violation_count) * 100, 2) if total_violation_count > 0 else 0

            conn.close()

            return {
                "status": True, "message": "success",
                "data": {"shortage_percentage": shortage_percentage}
            }

        except Exception as e:
            return {"status": False, "message": str(e), "data": {}}
    
    

    @staticmethod
    async def tibco_connection():
        try:
            creds = credential_loader.get_credentials('TIBCO')
            print("creds --->", creds)
            
            params = {
                "host": creds['host'],
                "database": creds['database'],
                "user": creds['user'],
                "password": creds['password'],
                "port": creds['port']
            }
            
            conn = mysql.connector.connect(**params)
            return conn
        except Exception as e:
            print(f"DB connection failed: {e}")
            return None
    
    @staticmethod
    async def execute_tibco_query(conn, query):
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(query)
            rows = cursor.fetchall()
            cursor.close()
            return {"data": rows}
        except mysql.connector.Error as e:
            print("Query execution failed:", e)
            return {"data": []}
        
    @staticmethod
    async def integrate_shortage_trips(filters, cross_filters, drill_state, payload):
        
        # ----- 1. Filter Separation and Date Condition Preparation ----- 
        trips_filters = filters  # All filters apply to trips
        
        # Extract Transporter Filter for later Pandas application (must be done post-merge)
        transporter_filter = next((f for f in trips_filters if getattr(f, 'key') == 'transporter_name'), None)
        trips_query = f"""
            SELECT *     
            FROM 
                sales_trips_till_date T
            WHERE load_status in ('6', '7')
        """
        sql_filters = [f for f in filters if getattr(f, "key", None) != "transporter_name"]    
        conditions = VTSAnalyticsActions.build_filter_conditions(sql_filters, cross_filters, trips_query)
        trips_query = VTSAnalyticsActions.apply_conditions_to_query(trips_query, conditions)
        print("trips_query", trips_query)
        trips_df = await VTSAnalyticsActions.execute_query(trips_query)
        if trips_df.empty:
            return {"status": "success", "total_invoice_count": 0, "total_vehicle_count": 0,
                    "filtered_invoice_count": 0, "filtered_vehicle_count": 0, "zones": []}
            
        trips_df.columns = [c.lower() for c in trips_df.columns]
        # 4b. Merge email master
        email_query = "SELECT transporter_code, transporter_name FROM email_master"
        email_df = await VTSAnalyticsActions.execute_query(email_query)

        if not email_df.empty:
            email_df.columns = [c.lower() for c in email_df.columns]

            # Clean and normalize keys
            email_df['transporter_code'] = (
                email_df['transporter_code']
                .astype(str)
                .str.strip()
                .str.replace(r'^00', '', regex=True)
            )
            trips_df['carrier_no'] = (
                trips_df['carrier_no']
                .astype(str)
                .str.strip()
                .str.replace(r'^00', '', regex=True)
            )
            

            # Ensure transporter_code is unique
            email_df = email_df.drop_duplicates(subset=['transporter_code'])

            # SAFE mapping: no extra rows, just add transporter_name column
            email_map = email_df.set_index('transporter_code')['transporter_name']
            trips_df['transporter_name'] = trips_df['carrier_no'].map(email_map)

            # --- Debug export for missing transporter_name ---
            # trips_df.to_csv('/Users/algofusion/Downloads/missing_transporters.csv', index=False)

        # ----- 5. Filter valid trips (Original Logic) -----
        
        # trips_df['qty_shortage'] = pd.to_numeric(trips_df['qty_shortage'], errors='coerce')
        trips_df['qty_shortage'] = (
            trips_df['qty_shortage']
            .astype(str)
            .str.replace(r"[^0-9.]", "", regex=True)  # keep only digits & decimal
            .str.strip()
        )

        trips_df['qty_shortage'] = pd.to_numeric(trips_df['qty_shortage'], errors='coerce').fillna(0)

        filtered_trips_df = trips_df[
            # trips_df['transporter_name'].notnull() &
            # trips_df['transporter_code'].notnull() &
            (trips_df['qty_shortage'] > 0)
        ].copy()

        # ----- 6. Apply Transporter Filter (Pandas, Original Logic) -----
        
        if transporter_filter:
            key = getattr(transporter_filter, "key", None)
            val = getattr(transporter_filter, "value", None)
            cond = getattr(transporter_filter, "cond", None)
            
            if key and val and key.lower() == 'transporter_name':
                df_col = key.lower()
                if cond == "equals":
                    filtered_trips_df = filtered_trips_df[filtered_trips_df[df_col] == val]
                elif cond == "in":
                    if not isinstance(val, list):
                        val = [val]
                    filtered_trips_df = filtered_trips_df[filtered_trips_df[df_col].isin(val)]

        # ----- 7. Counts after filtering (Original Logic) -----
        filtered_vehicle_count = filtered_trips_df['vehicle_id'].nunique()
        filtered_invoice_count = filtered_trips_df['invoice_no'].nunique()

        # filtered_vehicle_count = len(filtered_trips_df['vehicle_id'])
        
        if filtered_trips_df.empty:
            return {"status": "success", "total_invoice_count": 0, "total_vehicle_count": 0,
                    "filtered_invoice_count": 0, "filtered_vehicle_count": 0, "zones": []}
                    
        # filtered_trips_df = filtered_trips_df.drop_duplicates()

        # ----- 8. Convert invoice_date to IST (CRITICAL FIX: Original Logic) -----
        
        ist = pytz.timezone("Asia/Kolkata")
        if 'invoice_date' in filtered_trips_df.columns:
            filtered_trips_df['invoice_date'] = pd.to_datetime(filtered_trips_df['invoice_date'])
            
            if filtered_trips_df['invoice_date'].dt.tz is None:
                filtered_trips_df['invoice_date'] = filtered_trips_df['invoice_date'].dt.tz_localize('UTC').dt.tz_convert(ist)
            else:
                filtered_trips_df['invoice_date'] = filtered_trips_df['invoice_date'].dt.tz_convert(ist)
                
            filtered_trips_df['invoice_date'] = filtered_trips_df['invoice_date'].dt.strftime("%Y-%m-%d %H:%M:%S%z")

        # ----- 9. Dynamic hierarchical grouping (Original Logic) -----
        
        def compute_group_summary(df, group_cols):
            if not group_cols:
                return None

            result = []
            current_col = group_cols[0]
            next_cols = group_cols[1:]

            for keys, group in df.groupby(current_col, dropna=False):
                item = {current_col: keys}
                # item["shortage"] = group["qty_shortage"].sum()
                item["shortage"] = group["qty_shortage"].astype(float).sum()

                item["invoice_count"] = group["invoice_no"].nunique()
                # item["vehicle_count"] = group["vehicle_id"].nunique()
                # item["invoice_count"] = len(group["invoice_no"])
                # print('item["invoice_count"]', item["invoice_count"])
                item["vehicle_count"] = group["vehicle_id"].nunique()  
                

                # --- Material Group Bifurcation Logic ---
                if "material_group_nm" in group.columns and "qty_shortage" in group.columns:
                    bif_df = (
                        group
                        .groupby("material_group_nm", dropna=False)["qty_shortage"]
                        .sum()
                        .reset_index()
                    )

                    item["item_bifurcation"] = [
                        {
                            "material_group_nm": row["material_group_nm"],
                            "shortage": round(float(row["qty_shortage"]), 2),
                        }
                        for _, row in bif_df.iterrows()
                    ]


                child = compute_group_summary(group, next_cols)
                if child:
                    if next_cols[0] == "plant_nm":
                        item["plants"] = child
                    elif next_cols[0] == "transporter_name":
                        item["transporters"] = child
                    elif next_cols[0] == "vehicle_id":
                        item["vehicles"] = child
                    elif next_cols[0] == "invoice_no":
                        item["invoices"] = child
                else:
                    if current_col == "invoice_no" and "invoice_date" in group.columns:
                        item["invoice_date"] = group["invoice_date"].iloc[0]
                    

                result.append(item)

            return result

        filter_keys = [getattr(f, "key", None) for f in filters] if filters else []
        if "vehicle_id" in filter_keys:
            group_cols = ["vehicle_id", "invoice_no"]
        elif "transporter_name" in filter_keys:
            group_cols = ["transporter_name", "vehicle_id"]
        elif "plant_nm" in filter_keys:
            group_cols = ["plant_nm", "transporter_name"]
        elif "zone_nm" in filter_keys:
            group_cols = ["zone_nm", "plant_nm"]
        else:
            group_cols = ["zone_nm"]
        if "material_group_nm" not in filtered_trips_df.columns and "item_no" in filtered_trips_df.columns:
            filtered_trips_df.rename(columns={"item_no": "material_group_nm"}, inplace=True)
        date_wise = payload.get('date_wise')
        if date_wise is True or date_wise == "true":
            filtered_trips_df['created_at'] = pd.to_datetime(
                filtered_trips_df['invoice_date']
            ).dt.strftime("%Y-%m-%d")

            date_wise_df = (
                filtered_trips_df
                .groupby('created_at', as_index=False)
                .agg(
                    invoice_count=('invoice_no', 'nunique'),
                    vehicle_count=('vehicle_id', 'nunique'),
                    shortage=('qty_shortage', 'sum')
                )
                .sort_values('created_at')
            )
            date_wise_df['shortage'] = date_wise_df['shortage'].round(2)

            return {
                "status": "success",
                "message": "Date-wise data fetched successfully",
                "filtered_invoice_count": filtered_invoice_count,
                "filtered_vehicle_count": filtered_vehicle_count,
                "data": date_wise_df.to_dict(orient='records')
            }
        
        
        if payload.get('table') == "true":
            filtered_trips_df = filtered_trips_df.rename(
                columns={'material_group_nm': 'product_bifurcation', 'qty_shortage': 'shortage'}
            )
            filtered_trips_df['product_bifurcation'] = (
                filtered_trips_df['product_bifurcation'].astype(str)
                + ':' + filtered_trips_df['shortage'].astype(str)
            )

            table_df = (
                filtered_trips_df
                .groupby(['vehicle_id', 'invoice_no'], as_index=False)
                .agg({
                    'shortage': 'sum',  # sum shortages for same vehicle+invoice
                    'product_bifurcation': lambda x: ', '.join(x),
                    'plant_nm': 'first',
                    'zone_nm': 'first',
                    'transporter_name': 'first',
                    'invoice_date': 'first'
                })
            )
            


            if payload.get('date_wise') == "true":
                # Normalize invoice_date to date only (strip time)
                filtered_trips_df['created_at'] = pd.to_datetime(
                    filtered_trips_df['invoice_date']
                ).dt.strftime("%Y-%m-%d")

                date_wise_df = (
                    filtered_trips_df
                    .groupby('created_at', as_index=False)
                    .agg(
                        invoice_count=('invoice_no', 'nunique'),
                        vehicle_count=('vehicle_id', 'nunique'),
                        shortage=('qty_shortage', 'sum')
                    )
                    .sort_values('created_at')
                )

                date_wise_df['shortage'] = date_wise_df['shortage'].round(2)

                return {
                    "status": "success",
                    "message": "Date-wise data fetched successfully",
                    "filtered_invoice_count": filtered_invoice_count,
                    "filtered_vehicle_count": filtered_vehicle_count,
                    "data": date_wise_df.to_dict(orient='records')
                }
            # ---------- SHORTAGE FILTER ----------
            
            shortage_filter = payload.get("shortage_filter") # <=
            if shortage_filter:
                sf = str(shortage_filter).replace(" ", "")

                if "<" in sf:
                    limit = float(sf.split("<")[1])
                    table_df = table_df[table_df["shortage"] < limit]
                elif ">=" in sf or "≥" in sf:
                    limit = float(sf.split(">=")[1]) if ">=" in sf else float(sf.replace("≥", ""))
                    table_df = table_df[table_df["shortage"] >= limit]

            total_records = len(table_df)
            total_shortage = table_df["shortage"].sum()

            # ---------- SEARCH FILTER ----------
            search_text = payload.get("search_text") or payload.get("search")
            print("RAW SEARCH TEXT FROM PAYLOAD:", repr(search_text))

            if search_text:
                search_text = str(search_text).strip()
                print("SEARCH TEXT AFTER STRIP:", repr(search_text))

                if search_text:
                    # Optional: search only in specific columns
                    search_cols = [
                        "vehicle_id","invoice_no","plant_nm",
                        "zone_nm","transporter_name","product_bifurcation","invoice_date","shortage"
                    ]
                    search_cols = [c for c in search_cols if c in table_df.columns]

                    if search_cols:
                        mask = table_df[search_cols].astype(str).apply(
                            lambda col: col.str.contains(search_text, case=False, na=False),
                            axis=0).any(axis=1)

                        table_df = table_df[mask]

      
            page = int(payload.get("page", 1))
            page_size = int(payload.get("page_size", 100))  # default 100

            if page_size <= 0:
                page_size = total_records

            start = (page - 1) * page_size
            end = page * page_size

            paged_df = table_df.iloc[start:end]

            return {
                "status": "success",
                "message": "Table Data fetched successfully",
                "data": await safe_json(paged_df),
                "page": page,"page_size": page_size,
                "total_records": total_records,"total_shortage": total_shortage,
            } 
        

        filtered_trips_df = filtered_trips_df.replace([float('inf'), float('-inf')], None)
        filtered_trips_df = filtered_trips_df.where(pd.notnull(filtered_trips_df), None)
        print("TOTAL SHORTAGE BEFORE GROUPING =", filtered_trips_df["qty_shortage"].astype(float).sum())


        zones_list = compute_group_summary(filtered_trips_df, group_cols)

        
        def clean_for_json(obj):
            if isinstance(obj, dict):
                return {k: clean_for_json(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [clean_for_json(v) for v in obj]
            elif obj is None:
                return ""        # convert None -> empty string
            elif isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
                return ""        # convert NaN or inf -> empty string
            return obj

        # Clean nested structure (zones_list)
        zones_list = clean_for_json(zones_list)
        # -------------------------------------------------------

        from fastapi.encoders import jsonable_encoder


        response_data = {
            "status": "success",
            "filtered_invoice_count": filtered_invoice_count,
            "filtered_vehicle_count": filtered_vehicle_count,
            "zones": zones_list
        }

        return JSONResponse(content=jsonable_encoder(response_data))

    
    @staticmethod
    async def get_unblock_ageing(filters, cross_filters, drill_state, payload):
        try:
            # Cross filters
            _filters, daterange = await generate_cross_filter(cross_filters)
            current_date = datetime.now().strftime("%Y-%m-%d")

            closed_query = vts_query.vts_query.get("closed_alerts")
            shortage_query = vts_query.vts_query.get("unblocked_tt_shortage")

            # Drill Down filters for closed_query
            closed_query = await get_drill_down_filter(filters, closed_query)

            access_filters = [
                dashboard_studio_model.WidgetFiltersCreate(**rec)
                for rec in await hpcl_ceg_model.LpgOperationsSummary.get_clause_conditions(formated=True)
            ]

            closed_query = await widget_actions.WidgetActions.apply_filter_drilldown(closed_query, access_filters, drill_state)

            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, shortage_query)
            shortage_query = VTSAnalyticsActions.apply_conditions_to_query(shortage_query, conditions)
            shortage_query += " GROUP BY vehicle_id"

            # Date condition only for closed_query
            clause = "WHERE" if "where" not in closed_query.lower() else "AND"
            if daterange:
                closed_query += f" {clause} created_at BETWEEN {daterange}"
            else:
                closed_query += f" {clause} CAST(created_at AS DATE) = '{current_date}'"
            print("Final Shortage Query:", shortage_query)
            # Execute queries
            resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=closed_query, limit=0)
            shortage_resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=shortage_query, limit=0)

            # DataFrames
            df = pd.DataFrame(resp.get("data", []))
            

            df = await filter_data(df, _filters)

            shortage = pd.DataFrame(shortage_resp.get("data", []))

            # AVERAGE UNBLOCKING LOGIC (NEW)
            # Use blocked → unblocked dates from query
            df["vehicle_blocked_end_date"] = pd.to_datetime(df["vehicle_blocked_end_date"]).dt.tz_localize(None)
            df["vehicle_blocked_start_date"] = pd.to_datetime(df["vehicle_blocked_start_date"]).dt.tz_localize(None)
            
            df["created_at"] = pd.to_datetime(
                df["created_at"]
            ).dt.tz_localize(None)

            # consider only unblocked records
            df = df[df["vehicle_unblocked_date"].notna()]


            # Average Unblocking duration (days)
            df["unblocking_days"] = (
                (df["vehicle_unblocked_date"] - df["created_at"])
                .dt.total_seconds() / 86400
            ).clip(lower=0)
            # Violation counts
            violation_counts = (
                df.pivot_table(
                    index=["sap_id", "zone", "tt_number"],
                    columns="violation_type",
                    values="location_name",
                    aggfunc="count",
                    fill_value=0
                )
            )

            # Average Unblocking (group level)
            avg_unblocking = (
                df.groupby(
                    ["sap_id", "location_name", "transporter_code", "zone", "tt_number"],
                    as_index=False
                )
                .agg(
                    total_unblocking_days=("unblocking_days", "sum"),
                    total_alerts=("unblocking_days", "count")
                )
            )

            avg_unblocking["average_unblocking"] = (
                avg_unblocking["total_unblocking_days"]
                / avg_unblocking["total_alerts"]
            ).round(2)

            df = avg_unblocking.merge(
                violation_counts, on=["sap_id", "tt_number"], how="left"
            )

            df.columns.name = None

            # Merge shortage
            df = pd.merge(df, shortage, on="tt_number", how="left")
            df = df.fillna(0)

            # Ensure violation columns
            for col in [
                "continuous_driving_count", "device_tamper_count",
                "main_supply_removal_count", "night_driving_count",
                "route_deviation_count", "speed_violation_count",
                "stoppage_violations_count"
            ]:
                if col not in df.columns:
                    df[col] = 0

            df.rename(
                columns={
                    "continuous_driving_count": "CD",
                    "device_tamper_count": "DT",
                    "main_supply_removal_count": "PD",
                    "night_driving_count": "ND",
                    "route_deviation_count": "RD",
                    "speed_violation_count": "SV",
                    "stoppage_violations_count": "US"
                },
                inplace=True
            )

            # Drill-down aggregation
            if drill_state:
                group_by_keys = [drill_state]

                if filters:
                    filter_keys = [rec.key.strip('"') for rec in filters]

                    if "zone" in filter_keys and "location_name" not in filter_keys:
                        group_by_keys = ["zone", "location_name"]
                    elif (
                        "zone" in filter_keys and
                        "location_name" in filter_keys and
                        "transporter_code" not in filter_keys
                    ):
                        group_by_keys = ["zone", "location_name", "transporter_code"]
                    elif (
                        "zone" in filter_keys and
                        "location_name" in filter_keys and
                        "transporter_code" in filter_keys and
                        "tt_number" not in filter_keys
                    ):
                        group_by_keys = [
                            "zone", "location_name", "transporter_code", "tt_number"
                        ]

                df = df.groupby(group_by_keys, as_index=False).agg({
                    "CD": "sum",
                    "DT": "sum",
                    "PD": "sum",
                    "ND": "sum",
                    "RD": "sum",
                    "SV": "sum",
                    "US": "sum",
                    "total_alerts": "sum",
                    "total_unblocking_days": "sum",
                    "average_unblocking": "mean",
                    "shortage": "sum"
                })
                df["average_closing"] = (
                    df["total_alerts"] / df["total_unblocking_days"]
                ).round(2)
            
            pl_df = pl.from_pandas(df)
            if str(payload.get("download", "")).lower() == "true":
                return await download_streaming_data(pl_df, filename="unblock_ageing_data")

            return {
                "status": True,
                "message": "success",
                "data": df.to_dict(orient="records")
            }

        except Exception as e:
            print("-- Exception in get unblock ageing widget --")
            print("traceback :", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    
    async def get_emlock_open_data(filters, cross_filters, drill_state, payload):
        """
        Retrieve and process emlock open data with filters and drill-down (Polars).
        Supports date-wise aggregation when payload.date_wise = true
        """
        try:
            KEY_MAP = {
                "zone": "zone",
                "region": "region",
                "location": "location_name",
                "location_name": "location_name",
                "truck": "trucknumber",
                "trucknumber": "trucknumber",
                "invoice": "invoice_number",
                "invoice_number": "invoice_number",
            }

            def normalize_filters(filter_list):
                out = []
                for f in filter_list or []:
                    k = str(f.key).lower().strip()
                    if k in KEY_MAP:
                        f.key = KEY_MAP[k]
                    out.append(f)
                return out

            filters = normalize_filters(filters)

            # CROSS FILTERS (DATE RANGE)

            _filters, daterange = await generate_cross_filter(cross_filters)
            _filters = normalize_filters(_filters)

            current_date = datetime.now().strftime("%Y-%m-%d")

            #  BUILD SQL QUERY
            
            query = vts_query.vts_query.get("get_emlock_open_data")
            query = await get_drill_down_filter(filters, query)

            access_filters = [
                dashboard_studio_model.WidgetFiltersCreate(**rec)
                for rec in await hpcl_ceg_model.LpgOperationsSummary.get_clause_conditions(
                    formated=True
                )
            ]

            query = await widget_actions.WidgetActions.apply_filter_drilldown(
                query, access_filters, drill_state
            )

            clause = "WHERE" if "where" not in query.lower() else "AND"
            query += (
                f" {clause} createdat BETWEEN {daterange}"
                if daterange
                else f" {clause} CAST(createdat AS DATE) = '{current_date}'"
            )

            # FETCH DATA
            resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=query, limit=0)
            df = pl.DataFrame(resp.get("data", []))

            if df.is_empty():
                return {"status": True, "message": "No data found", "data": []}

            # ADD EVENT DATE (FOR DATE-WISE)

            if "createdat" in df.columns:
                df = df.with_columns(
                    pl.col("createdat").cast(pl.Date).alias("event_date")
                )

            # Apply cross filters (BU / ZONE / SAP / etc)
            df = await filter_data(df, _filters)

            # Deduplication
            dedup_cols = [
                c for c in ["invoice_number", "trucknumber", "zone", "location_name"]
                if c in df.columns
            ]
            if dedup_cols:
                df = df.unique(subset=dedup_cols, keep="first")

            #  COMPLETED INVOICE HELPERS

            def apply_status_filter(pl_df, status):
                if not status:
                    return pl_df

                status = status.lower().strip()

                if "trip_status" in pl_df.columns:
                    if status == "live":
                        return pl_df.filter(pl.col("trip_status") != "Closed")
                    if status == "closed":
                        return pl_df.filter(pl.col("trip_status") == "Closed")

                return pl_df

            def add_swipe_cols(pl_df):
                return pl_df.with_columns([
                    pl.col("swipeoutl1")
                    .fill_null("")
                    .cast(pl.Utf8)
                    .str.to_lowercase()
                    .eq("false")
                    .alias("has_swipeoutl1"),

                    pl.col("swipeoutl2")
                    .fill_null("")
                    .cast(pl.Utf8)
                    .str.to_lowercase()
                    .eq("false")
                    .alias("has_swipeoutl2"),
                ])

            #  MODES: DOWNLOAD / SEARCH / TABLE
            if payload.get("download", "").lower() == "true":
                df = apply_status_filter(df, payload.get("status"))
                df = add_swipe_cols(df)
                df = df.filter(pl.col("has_swipeoutl1") | pl.col("has_swipeoutl2"))
                df = df.drop(["has_swipeoutl1", "has_swipeoutl2"], strict=False)
                return await download_streaming_data(df, filename="emlock_open_data")

            if payload.get("search") == "true":
                return {"status": True, "message": "success", "data": df.to_dicts()}

            if payload.get("table") == "true":
                df = apply_status_filter(df, payload.get("status"))
                df = add_swipe_cols(df)
                df = df.filter(pl.col("has_swipeoutl1") | pl.col("has_swipeoutl2"))
                return {
                    "status": True,
                    "message": "success",
                    "data": df.to_dicts(),
                    "total_records": df.height,
                }

            
            # AGGREGATION
            df = add_swipe_cols(df)

            base_cols = [
                c for c in
                ["zone", "region", "location_name", "invoice_number", "trucknumber", "event_date"]
                if c in df.columns
            ]

            base = (
                df.group_by(base_cols)
                .agg([
                    pl.any("has_swipeoutl1").alias("has_swipeoutl1"),
                    pl.any("has_swipeoutl2").alias("has_swipeoutl2"),
                ])
            )

            base = apply_status_filter(base, payload.get("status"))

            filtered = base.filter(
                pl.col("has_swipeoutl1") | pl.col("has_swipeoutl2")
            )

            #  DATE-WISE OVERRIDE
            if payload.get("date_wise") is True:
                group_by = ["event_date"]
            else:
                group_by = ["zone"]
                keys = [f.key for f in filters or []]

                if "zone" in keys and "location_name" not in keys:
                    group_by = ["location_name"]
                if "location_name" in keys and "trucknumber" not in keys:
                    group_by = ["trucknumber"]
                if "trucknumber" in keys:
                    group_by = ["invoice_number"]

            group_by = [g for g in group_by if g in filtered.columns]

            grouped = (
                filtered.group_by(group_by)
                .agg([
                    pl.col("invoice_number")
                    .filter(pl.col("has_swipeoutl1"))
                    .n_unique()
                    .alias("swipeoutl1_count"),

                    pl.col("invoice_number")
                    .filter(pl.col("has_swipeoutl2"))
                    .n_unique()
                    .alias("swipeoutl2_count"),

                    pl.col("invoice_number").n_unique().alias("distinct_invoice_count"),
                    pl.col("trucknumber").n_unique().alias("distinct_vehicle_count"),
                ])
            )

            return {
                "status": True,
                "message": "success",
                "swipe_out_l1_count": int(
                    filtered.filter(pl.col("has_swipeoutl1"))["invoice_number"].n_unique()
                ),
                "swipe_out_l2_count": int(
                    filtered.filter(pl.col("has_swipeoutl2"))["invoice_number"].n_unique()
                ),
                "distinct_invoice_count": int(filtered["invoice_number"].n_unique()),
                "distinct_vehicle_count": int(filtered["trucknumber"].n_unique()),
                "data": grouped.to_dicts(),
            }

        except Exception:
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": "Internal error", "data": []}
        
    @staticmethod
    async def power_disconnection(filters, cross_filters, drill_state, payload):
        try:
            #  Get base query and apply filters
            query = vts_query.vts_query.get(drill_state.split(",")[0])
            if not query:
                return {"status": False, "message": "Query not found", "data": []}
            
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, query)
            final_query = VTSAnalyticsActions.apply_conditions_to_query(query, conditions)
            vts_df = await VTSAnalyticsActions.execute_query(final_query)
            vts_df = vts_df.drop_duplicates(subset=['invoice_number'], keep='first')
            
            if vts_df.empty:
                return {"status": True, "message": "No power disconnection alerts found", "data": []}

            trans_query = """SELECT truck_no, transporter_name from vts_truck_master"""
            df_transporter = await VTSAnalyticsActions.execute_query(trans_query)
            
            if df_transporter.empty:
                return {"status": False, "message": "No matching vehicle details found in alerts", "data": []}
            
            merged_df = vts_df.merge(df_transporter, left_on="tl_number", right_on="truck_no", how="left")
            
            if merged_df.empty:
                return {"status": False, "message": "No valid zone data found after merging with alerts", "data": []}
           
            # Filter for power disconnection violations (>= 6)
            violation_type = "main_supply_removal_count"
            violation_filtered_df = merged_df[merged_df[violation_type].fillna(0) >= 6].copy()
            
            # Remove empty values for zone, location, transporter
            for key in ["zone", "location_name", "transporter_name"]:
                if payload.get(key):
                    violation_filtered_df = violation_filtered_df[violation_filtered_df[key] == payload[key]]
           # --- DATE-WISE AGGREGATION (move here!) ---
            date_wise = payload.get("date_wise") or payload.get("payload", {}).get("date_wise")
            if date_wise is True or date_wise == "true":
                violation_filtered_df['created_at'] = pd.to_datetime(violation_filtered_df['created_at'])
                ist = pytz.timezone("Asia/Kolkata")
                if violation_filtered_df['created_at'].dt.tz is None:
                    violation_filtered_df['created_at'] = violation_filtered_df['created_at'].dt.tz_localize('UTC').dt.tz_convert(ist)
                else:
                    violation_filtered_df['created_at'] = violation_filtered_df['created_at'].dt.tz_convert(ist)
                
                violation_filtered_df['date'] = violation_filtered_df['created_at'].dt.strftime("%Y-%m-%d")
                
                date_wise_df = (
                    violation_filtered_df.groupby('date')
                    .agg(
                        invoice_count=('invoice_number', 'nunique'),
                        violation_count_more_than_6=(violation_type, 'count'),
                        total_violations=(violation_type, 'sum'),
                        vehicle_count=('tl_number', 'nunique')
                    )
                    .reset_index()
                    .sort_values('date')
                )
                
                return {
                    "status": True,
                    "message": f"{violation_type} date-wise data",
                    "data": date_wise_df.to_dict(orient='records')
                } 
            if violation_filtered_df.empty:
                return {"status": True, "message": "No data found for the applied filters", "data": []}
            
            #  TL-level drill-down for invoice details
            selected_tl = payload.get("tl_number")
            if selected_tl:
                violation_filtered_df = violation_filtered_df[violation_filtered_df["tl_number"] == selected_tl]
                
                if violation_filtered_df.empty:
                    return {"status": True, "message": f"No invoices found for vehicle {selected_tl}", "data": []}
                
                # Return invoice details sorted by created_at
                invoice_df = violation_filtered_df.sort_values(by="created_at", ascending=True)
                invoice_df = invoice_df[["invoice_number", "created_at", violation_type]]
                
                # Rename columns for frontend
                invoice_df.rename(columns={
                    "invoice_number": "invoice_no",
                    "created_at": "created_at"
                }, inplace=True)
                
                result = invoice_df.to_dict(orient="records")
                return {"status": True, "message": f"{violation_type} details for vehicle {selected_tl}", "data": result}
            
            # Determine grouping column for summaries
            if payload.get("transporter_name"):
                group_col = "tl_number"
            elif payload.get("location_name"):
                group_col = "transporter_name"
            elif payload.get("zone"):
                group_col = "location_name"
            else:
                group_col = "zone"
            
            # Summarize counts
            # violation_count_more_than_6: Count of invoices where main_supply_removal_count >= 6
            # total_violations: Sum of actual main_supply_removal_count values
            summary_df = (
                violation_filtered_df.groupby(group_col)
                .agg({
                    "invoice_number": pd.Series.nunique,  # invoice_count
                    violation_type: ['count', 'sum']  # count of records >= 6, and sum of actual values
                })
                .reset_index()
            )
            
            # Flatten multi-level columns
            summary_df.columns = [group_col, 'invoice_count', 'violation_count_more_than_6', 'total_violations']
            
            if group_col != "tl_number":
                summary_df["vehicle_count"] = violation_filtered_df.groupby(group_col)["tl_number"].nunique().values
            
            result = summary_df.to_dict(orient="records")
            return {"status": True, "message": f"{violation_type} drill-down data", "data": result}
        
        except Exception as e:
            print("Exception in power_disconnection:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": []}
    

    @staticmethod
    async def risk_score(filters, cross_filters, drill_state, payload):
        """
        Fetch paginated data from the specified risk score table and also support downloading.
        """
        try:
            if hasattr(payload, "dict"):
                payload = payload.dict()
            elif not isinstance(payload, dict):
                payload = dict(payload) if hasattr(payload, "__iter__") else {}
            
            table_name = payload.get("table_name")
            columns = payload.get("columns")
            limit = 0 if payload.get("download") == "true" else payload.get("page_size", 100)
            conditions = []
            # Pagination parameters from payload
            page = int(payload.get("page", 0))

            if not table_name:
                return {"status": False, "message": "table_name not provided in payload", "data": []}

            print(f"Fetching data from table: {table_name}")

            # ==================== HANDLE RISK_SCORE_TRENDS ====================
            action = payload.get("action")
            if action == "risk_score_trends":
                TRENDS_CONFIG = {
                    "tt_risk_score": {
                        "payload_key": "clicked_tt_number",
                        "query_key": "tt_risk_score_daily_violations",
                        "message_template": "Violation trends for TT {}"
                    },
                    "transporter_risk_score": {
                        "payload_key": "clicked_transporter_code",
                        "query_key": "transporter_risk_score_daily_violations",
                        "message_template": "Violation trends for transporter {}"
                    }
                }

                if table_name not in TRENDS_CONFIG:
                    return {"status": False, "message": f"Table {table_name} not supported for trends", "data": []}

                config = TRENDS_CONFIG[table_name]
                clicked_value = payload.get(config["payload_key"])

                if not clicked_value:
                    return {"status": False, "message": f"{config['payload_key']} not provided in payload", "data": []}

                safe_value = str(clicked_value).replace("'", "''")
                violation_query_template = vts_query.vts_query.get(config["query_key"])

                if not violation_query_template:
                    return {"status": False, "message": f"Query template not found for {config['query_key']}", "data": []}

                violation_query = violation_query_template.format(safe_value)
                print(f"Executing trends query: {violation_query}")

                response = await urdhva_base.BasePostgresModel.get_aggr_data(query=violation_query, limit=0, skip_total=True)
                violation_data = response.get('data', [])

                # ========== VIOLATION GROUPING LOGIC WITH POLARS ==========
                if violation_data:
                    df = pl.DataFrame(violation_data, infer_schema_length=None)
                    violation_columns_map = vts_query.vts_query.get("violation_columns_map", {})

                    # Convert violation_date to date format and format for display
                    df = df.with_columns(
                        pl.col('violation_date').cast(pl.Date).cast(pl.Utf8).str.replace_all(
                            r'(\d{4})-(\d{2})-(\d{2})', r'$2-$3-$1'
                        ).alias('violation_date_formatted')
                    )

                    # Get all violation type columns that exist
                    violation_types = [col for col in violation_columns_map.keys() if col in df.columns]

                    # Group by date and sum violations
                    agg_cols = [pl.col(vtype).sum().cast(pl.Int64) for vtype in violation_types]
                    if 'risk_score' in df.columns:
                        agg_cols.append(pl.col('risk_score').first())

                    grouped_df = df.group_by('violation_date_formatted').agg(agg_cols).sort('violation_date_formatted')

                    # Transform to nested structure
                    grouped_data = []
                    for row in grouped_df.iter_rows(named=True):
                        date = row.pop('violation_date_formatted')
                        risk_score = row.pop('risk_score', None)
                        violations = [
                            {"violation_type": k, "count": int(row[k]) if row[k] is not None else 0}
                            for k in violation_types
                        ]
                        violations.sort(key=lambda x: x['violation_type'])
                        
                        record_entry = {"date": date, "records": violations}
                        if risk_score is not None:
                            record_entry["risk_score"] = float(risk_score) if risk_score is not None else 0.0
                            
                        grouped_data.append(record_entry)
                else:
                    grouped_data = []

                if not grouped_data:
                    return {"status": True, "message": f"No violation trends found for {clicked_value}", "data": []}

                return {
                    "status": True,
                    "message": config["message_template"].format(clicked_value),
                    "data": grouped_data,
                    "total_records": len(grouped_data)
                }

            # ==================== HANDLE CLICK HANDLERS ====================

            CLICK_HANDLERS = {
                "completed_trips_risk_score": {
                    "payload_key": "clicked_invoice_no",
                    "table": "public.combo_alerts",
                    "column": "invoice_no",
                    "message_template": "Combo alerts for invoice {}"
                },
                "cluster_master": {
                    "payload_key": "clicked_cluster_id",
                    "table": "public.clusterwise_event",
                    "column": "cluster_id",
                    "message_template": "Cluster events for cluster_id {}"
                },
                "transporter_risk_score": {
                    "payload_key": "clicked_transporter_code",
                    "table": "public.transporter_risk_score",
                    "column": "transporter_code",
                    "message_template": "Transporter events for transporter_code {}"
                },
                "tt_risk_score": {
                    "payload_key": "clicked_tt_number",
                    "table": "public.tt_risk_score",
                    "column": "tt_number",
                    "message_template": "Transporter events for tt_number {}"
                }
            }
            if table_name in CLICK_HANDLERS:
                config = CLICK_HANDLERS[table_name]
                clicked_value = payload.get(config["payload_key"])
                
                if clicked_value:
                    safe_value = str(clicked_value).replace("'", "''")
                    query = f"SELECT * FROM {config['table']} WHERE {config['column']} = '{safe_value}'"
                    
                    response = await urdhva_base.BasePostgresModel.get_aggr_data(query=query, limit=0, skip_total=True)
                    data = response.get('data', [])
                    
                    # Handle distinct location\_name grouping or specific location filtering
                    loc_val = payload.get("location_name")
                    if loc_val and data:
                        if loc_val is True:
                            df = pl.DataFrame(data, infer_schema_length=None)
                            if "location_name" in df.columns:
                                grouped = df.group_by("location_name").agg(pl.len().alias("count"))
                                data = grouped.to_dicts()
                        elif isinstance(loc_val, str):
                            data = [row for row in data if row.get("location_name") == loc_val]

                    # Check if download is requested
                    if payload.get("download") == "true":
                        df = pd.DataFrame(data)
                        if not df.empty:
                            # Remove timezone info from datetime columns
                            for col in df.select_dtypes(include=["datetime64[ns, UTC]", "datetimetz"]).columns:
                                df[col] = df[col].dt.tz_localize(None)
                            
                            pl_df = pl.from_pandas(df)
                            return await download_streaming_data(pl_df, filename=f'{table_name}_{clicked_value}')
                    
                    return {
                        "status": True,
                        "message": config["message_template"].format(clicked_value),
                        "data": data,
                        "total_records": len(data)
                    }

            if columns and isinstance(columns, list) and "zone" in columns:
                columns.remove("zone")

            if columns and isinstance(columns, list) and columns:
                select_columns = ", ".join([f'"{col}"' for col in columns])
                base_query = f'SELECT {select_columns} FROM public."{table_name}"'
            else:
                base_query = f'SELECT * FROM public."{table_name}"'

            access_filters = [
                dashboard_studio_model.WidgetFiltersCreate(**rec)
                for rec in await hpcl_ceg_model.LpgOperationsSummary
                .get_clause_conditions(formated=True)
            ]
            base_query = await widget_actions.WidgetActions.apply_filter_drilldown(
                base_query, access_filters, drill_state
            )

            # Build and apply conditions
            conditions = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)

            # Add version_date filter if a specific version date is provided
            vd = payload.get("version_date")
            if vd and not any("version_date" in str(c).lower() for c in conditions):
                conditions.append(f"version_date::date = '{vd}'")

            # Add search filter condition if a search term is provided
            search_term = payload.get("search")
            if search_term and columns:
                search_conditions = []
                for col in columns:
                    search_conditions.append(f'CAST("{col}" AS TEXT) ILIKE \'%{search_term}%\'')
                if search_conditions:
                    conditions.append(f"({' OR '.join(search_conditions)})")

            # Add column-specific search filters
            column_filters = payload.get("column_filters")
            if column_filters and isinstance(column_filters, dict):
                for col, search_val in column_filters.items():
                    if search_val:  # Ensure there is a value to search for
                        # Add a case-insensitive search condition for the specific column
                        conditions.append(f'CAST("{col}" AS TEXT) ILIKE \'%{search_val}%\'')
            
            # Add column-specific range filters (>=, <=)
            range_filters = payload.get("range_filters")
            if range_filters and isinstance(range_filters, list):
                for r_filter in range_filters:
                    col = r_filter.get("column")
                    op = r_filter.get("operator")
                    val = r_filter.get("value")

                    if col and op and val is not None:
                        supported_operators = ['>=', '<=', '>', '<', '=', '!=']
                        if op not in supported_operators:
                            continue

                        conditions.append(f'CAST("{col}" AS NUMERIC) {op} {val}')

            # Skip COUNT query for downloads — wasteful round-trip on large tables
            is_download = payload.get("download") == "true"
            total_records = 0
            if not is_download:
                count_query = f'SELECT COUNT(*) FROM public."{table_name}"'
                filtered_count_query = VTSAnalyticsActions.apply_conditions_to_query(count_query, conditions)
                count_resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=filtered_count_query)
                total_records = count_resp['data'][0]['count'] if count_resp.get('data') else 0

                        # Add sorting
            sort_by = payload.get("sort_by")
            sort_direction = payload.get("sort_direction", "asc").upper()
            if sort_by and sort_direction in ["ASC", "DESC"]:
                # Add ORDER BY to the base query before applying other conditions
                # Note: apply_conditions_to_query handles placing this correctly
                base_query += f' ORDER BY "{sort_by}" {sort_direction}'

            # Build final filtered query
            # For completed_trips_risk_score download: query is used directly inside stream_csv()
            filtered_data_query = VTSAnalyticsActions.apply_conditions_to_query(base_query, conditions)
            resp = None
            if not (is_download and table_name == "completed_trips_risk_score"):
                resp = await urdhva_base.BasePostgresModel.get_aggr_data(
                    query=filtered_data_query, limit=limit, skip=page, skip_total=True
                )
                if not resp['data']:
                    return {"status": True, "message": "No data found", "data": [], "total_records": 0}

            if is_download:
                def norm_id(v):
                    try: return str(int(float(str(v).strip())))
                    except Exception: return str(v).strip()

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"{table_name}_{timestamp}.xlsx"
                output    = io.BytesIO()

                if table_name == "cluster_master":
                    # Fetch ALL cluster_master rows (no pagination limit)
                    def strip_tz(frame):
                        for c in frame.select_dtypes(include=["datetime64[ns, UTC]", "datetimetz"]).columns:
                            frame[c] = frame[c].dt.tz_localize(None)
                        return frame

                    all_master_df = strip_tz(pd.DataFrame(
                        (await urdhva_base.BasePostgresModel.get_aggr_data(
                            query=VTSAnalyticsActions.apply_conditions_to_query(base_query, conditions),
                            limit=0, skip_total=True
                        )).get("data", [])
                    )).dropna(axis=1, how="all")

                    cluster_ids = list(dict.fromkeys(
                        norm_id(v) for v in all_master_df.get("cluster_id", pd.Series()).dropna()
                    ))

                    # Fetch ALL clusterwise_event rows for these clusters + same date filter
                    event_df = pd.DataFrame()
                    if cluster_ids:
                        ids_sql        = ", ".join(f"'{c}'" for c in cluster_ids)
                        ev_base        = f"SELECT * FROM public.clusterwise_event WHERE cluster_id::text IN ({ids_sql})"
                        ev_conds       = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, ev_base)
                        vd             = payload.get("version_date", "")
                        if vd and not any("version_date" in c for c in ev_conds):
                            ev_conds.append(f"version_date::date = '{vd}'")
                        event_df = strip_tz(pd.DataFrame(
                            (await urdhva_base.BasePostgresModel.get_aggr_data(
                                query=VTSAnalyticsActions.apply_conditions_to_query(ev_base, ev_conds),
                                limit=0, skip_total=True
                            )).get("data", [])
                        ))
                        if not event_df.empty and "cluster_id" in event_df.columns:
                            event_df["cluster_id"] = event_df["cluster_id"].apply(norm_id)

                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        wb  = writer.book
                        lnk = wb.add_format({"font_color": "#1155CC", "underline": True})
                        bck = wb.add_format({"font_color": "#C00000", "underline": True, "bold": True, "bg_color": "#FFF2CC", "border": 1})
                        hdr = wb.add_format({"bold": True, "bg_color": "#D9E1F2", "border": 1})

                        # Summary sheet first → Tab #1, active on open
                        all_master_df.to_excel(writer, index=False, sheet_name="Summary")
                        sw  = writer.sheets["Summary"]
                        sw.activate()
                        cols = list(all_master_df.columns)
                        sheets_written = set()

                        if "cluster_id" in cols:
                            ci = cols.index("cluster_id")
                            sw.write(0, ci, "cluster_id", hdr)
                            for r, v in enumerate(all_master_df["cluster_id"], 1):
                                sn = norm_id(v)[:31]
                                sw.write_url(r, ci, f"internal:'{sn}'!A1", lnk, norm_id(v))

                        # Per-cluster sheets with ← Back to Summary
                        if not event_df.empty and "cluster_id" in event_df.columns:
                            for cid, sdf in event_df.groupby("cluster_id", sort=True):
                                sn = str(cid)[:31]
                                sdf.to_excel(writer, index=False, sheet_name=sn, startrow=1)
                                cws = writer.sheets[sn]
                                cws.write_url(0, 0, "internal:'Summary'!A1", bck, "← Back to Summary")
                                cws.set_column(0, 0, 22)
                                sheets_written.add(str(cid))

                        # Replace links for clusters with no events → plain text
                        if "cluster_id" in cols:
                            ci = cols.index("cluster_id")
                            for r, v in enumerate(all_master_df["cluster_id"], 1):
                                if norm_id(v) not in sheets_written:
                                    sw.write(r, ci, norm_id(v))

                # completed_trips_risk_score - CSV STREAMING
                # Large table: bypass Pandas/Polars entirely, stream row-by-row from DB to client.
                # Each row is yielded immediately - nginx never idles → no 60s timeout.
                elif table_name == "completed_trips_risk_score":
                    print(f"[download] Streaming {table_name} natively as CSV...")
                    from sqlalchemy import text
                    import urdhva_base.postgresmodel as pm

                    async def stream_csv():
                        import csv
                        session = await pm.manager.get_session()
                        try:
                            result = await session.stream(
                                text(filtered_data_query).execution_options(yield_per=1000)
                            )
                            cols_fetched = False
                            async for row in result:
                                output_buf = io.StringIO()
                                writer = csv.writer(output_buf)
                                if not cols_fetched:
                                    writer.writerow(list(result.keys()))
                                    cols_fetched = True
                                row_vals = []
                                for v in row:
                                    if hasattr(v, "tzinfo") and getattr(v, "tzinfo") is not None:
                                        v = v.replace(tzinfo=None).isoformat(sep=" ")
                                    elif getattr(type(v), "__name__", "") == "datetime":
                                        v = v.isoformat(sep=" ")
                                    row_vals.append(v)
                                writer.writerow(row_vals)
                                # Instantly flush each row to nginx & client — no idle timeout
                                yield output_buf.getvalue().encode('utf-8')
                        except Exception as e:
                            print(f"[download] Streaming error: {e}")
                        finally:
                            import asyncio
                            await asyncio.shield(session.close())

                    file_name = f"completed_trips_risk_score_{str(datetime.utcnow().date())}.csv"
                    return StreamingResponse(
                        stream_csv(),
                        media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="{file_name}"'}
                    )

                # BRANCH 3: all other tables - single-sheet XLSX (small data, BytesIO)
                else:
                    def strip_tz(frame):
                        for c in frame.select_dtypes(include=["datetime64[ns, UTC]", "datetimetz"]).columns:
                            frame[c] = frame[c].dt.tz_localize(None)
                        return frame

                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        strip_tz(pd.DataFrame(resp['data'])).dropna(axis=1, how="all") \
                            .to_excel(writer, index=False, sheet_name=table_name[:31])

                output.seek(0)
                headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=headers
                )
            
            return {
                "status": True,
                "message": f"Successfully fetched {len(resp['data'])} records from {table_name}",
                "data": resp['data'],
                "page": page,
                "page_size": limit,
                "total_records": total_records
            }
        except Exception as e:
            print("Exception in risk_score:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": [], "total_records": 0}
    @staticmethod
    async def risk_score_trends(filters, cross_filters, drill_state, payload):
        """
        Wrapper around risk_score that sets action to 'risk_score_trends'.
        This ensures the dispatcher can find it via hasattr and routes it
        through the standard risk_score logic for violation trends.
        """
        if not isinstance(payload, dict):
            payload = {}
        payload['action'] = 'risk_score_trends'
        return await VTSAnalyticsActions.risk_score(
            filters=filters,
            cross_filters=cross_filters,
            drill_state=drill_state,
            payload=payload
        )
    
    @staticmethod
    async def risk_score_cluster_map(filters, cross_filters, drill_state, payload):
        """
        Fetch cluster_master records and their events.
        If version_date is passed without cluster_id: Gives highly optimized summary map view + event frequency counts.
        If cluster_id is passed: Provides a specific deep-dive drilldown mapping exact events.
        If event_lat_lon is passed: Further filters the drill-down purely to events matching that coordinate.
        Applies dynamic global filters.
        """
        try:
            from collections import defaultdict
            if hasattr(payload, "dict"):
                payload = dict(payload)
            elif not isinstance(payload, dict):
                payload = dict(payload) if hasattr(payload, "__iter__") else {}
                
            cluster_id = payload.get("clicked_cluster_id") or payload.get("cluster_id")
            version_date = payload.get("version_date")
            event_lat_lon = payload.get("clicked_event_lat_lon") or payload.get("event_lat_lon")
            
            if not cluster_id and not version_date:
                return {"status": False, "message": "Either cluster_id or version_date must be provided in payload", "data": []}
                
            # Build general dynamic filters
            base_conds = []
            if filters:
                # Utilizing existing build_filter_conditions helper for dynamic filter mappings
                # (Assumes "SELECT * FROM public.cluster_master m" for correct DB parsing context)
                built_conds = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, "SELECT * FROM public.cluster_master")
                if isinstance(built_conds, list):
                    base_conds.extend(built_conds)
                elif built_conds:
                    base_conds.append(built_conds)
            
            # Additional structural logic
            filter_str = " AND ".join(base_conds)
            if filter_str:
                filter_str = f" AND {filter_str}" # Append logic

            if cluster_id:
                # SECOND/THIRD DRILLDOWN (Specific Cluster & Optional Coordinate)
                safe_cluster_id = str(cluster_id).replace("'", "''")
                master_cond = f"m.cluster_id::text = '{safe_cluster_id}'"
                if version_date:
                    safe_date = str(version_date).replace("'", "''")
                    master_cond += f" AND m.version_date::date = '{safe_date}'"
                    
                master_query = vts_query.vts_query.get("cluster_map_master_drilldown").format(
                    master_cond=master_cond, filter_str=filter_str
                )
                
                event_cond = f" AND e.event_lat_lon = '{str(event_lat_lon).replace(chr(39), chr(39)+chr(39))}'" if event_lat_lon else ""
                event_query = vts_query.vts_query.get("cluster_map_event_drilldown").format(
                    master_cond=master_cond, filter_str=filter_str, event_cond=event_cond
                )
                
                master_resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=master_query, limit=0, skip_total=True)
                master_data = master_resp.get('data', [])
                
                if not master_data:
                    msg = f"No master data found for cluster_id {cluster_id}"
                    return {"status": True, "message": msg, "data": []}
                    
                event_resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=event_query, limit=0, skip_total=True)
                event_data = event_resp.get('data', [])
                
                cluster_info = master_data[0]
                cluster_info['cluster_events'] = event_data
                
                resp_msg = f"Cluster coordinate events for {event_lat_lon}" if event_lat_lon else f"Cluster map details for cluster_id {cluster_id}"
                return {
                    "status": True,
                    "message": resp_msg,
                    "data": [cluster_info],
                    "total_records": 1
                }
            
            else:
                # FIRST DRILLDOWN (Fast summary of ALL Clusters based on Version Date)
                safe_date = str(version_date).replace("'", "''")
                
                # Fetch only summary master data
                master_query = vts_query.vts_query.get("cluster_map_master_summary").format(
                    safe_date=safe_date, filter_str=filter_str
                )
                
                # Aggregate events to count in the DB to avoid gigantic payloads, but grab their coordinates
                # Ensure the same filtering is applied backwards to the events through EXISTS so counts stay accurate
                count_query = vts_query.vts_query.get("cluster_map_count_summary").format(
                    safe_date=safe_date, filter_str=filter_str
                )
                
                master_resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=master_query, limit=0, skip_total=True)
                count_resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=count_query, limit=0, skip_total=True)
                
                master_data = master_resp.get('data', [])
                count_data = count_resp.get('data', [])
                
                # Create easy dictionary lookup
                counts_map = {str(c.get('c_id')): c.get('event_count', 0) for c in count_data if c.get('c_id')}
                coords_map = {str(c.get('c_id')): c.get('event_coords', []) for c in count_data if c.get('c_id')}
                
                total_cluster_events = 0
                risk_bands = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
                
                formatted_data = []
                for m in master_data:
                    c_id = str(m.get('cluster_id')) if m.get('cluster_id') is not None else None
                    ev_count = counts_map.get(c_id, 0)
                    ev_coords_raw = coords_map.get(c_id, [])
                    # Clean out nulls from aggregation
                    ev_coords = [c for c in ev_coords_raw if c]
                    
                    total_cluster_events += ev_count
                    
                    # Accumulate risk_bands
                    rb = str(m.get('risk_band', '')).strip().title()
                    if rb == 'Critical':
                        risk_bands['Critical'] += 1
                    elif rb == 'High':
                        risk_bands['High'] += 1
                    elif rb in ('Medium', 'Moderate'):
                        risk_bands['Medium'] += 1
                    elif rb == 'Low':
                        risk_bands['Low'] += 1
                    
                    # Ensure centroid parsing
                    cl_ll = m.get("centroid_lat_lon")
                    if not cl_ll and m.get("centroid_lat") and m.get("centroid_lon"):
                        cl_ll = f"{m.get('centroid_lat')},{m.get('centroid_lon')}"
                        
                    formatted_data.append({
                        "cluster_id": m.get("cluster_id"),
                        "cluster_lat_long": cl_ll,
                        "risk_score": m.get("risk_score"),
                        "cluster_events_count": ev_count,
                        "cluster_event_coordinates": ev_coords,
                        "risk_band": m.get("risk_band"),
                        "city": m.get("plant_name") or m.get("city"),
                        "state": m.get("state") or m.get("zone_name") or "",
                        "first_seen": m.get("first_seen"),
                        "last_seen": m.get("last_seen"),
                        "type": m.get("location_type"),
                        # Include raw master for complete metadata if ever needed by frontend
                        **m,
                    })

                return {
                    "status": True,
                    "message": f"All cluster map details for version_date {version_date}",
                    "data": formatted_data,
                    "total_clusters": len(formatted_data),
                    "total_cluster_events": total_cluster_events,
                    "High": risk_bands["High"] + risk_bands["Critical"], # Optionally combine depending on map thresholds
                    "Medium": risk_bands["Medium"],
                    "Low": risk_bands["Low"]
                }
        except Exception as e:
            print("Exception in risk_score_cluster_map:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": [], "total_records": 0}

    @staticmethod
    async def cluster_wise_daily_trends(filters, cross_filters, drill_state, payload):
        try:
            if hasattr(payload, "dict"):
                payload = payload.dict()
            elif not isinstance(payload, dict):
                payload = dict(payload) if hasattr(payload, "__iter__") else {}

            cluster_id = payload.get("cluster_id") or payload.get("clicked_cluster_id")
            
            filter_sql = ""
            if cluster_id:
                safe_val = str(cluster_id).replace("'", "''")
                filter_sql = f" AND cluster_id::text = '{safe_val}'"

            base_conds = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, "SELECT * FROM public.clusterwise_event")
            if base_conds:
                if isinstance(base_conds, list):
                    filter_sql += " AND " + " AND ".join(base_conds)
                else:
                    filter_sql += f" AND {base_conds}"

            query = vts_query.vts_query.get("cluster_wise_daily_trends").format(filter_sql=filter_sql)
            resp = await urdhva_base.BasePostgresModel.get_aggr_data(query=query, skip_total=True)
            
            data = resp.get('data', [])
            for row in data:
                if 'day' in row and row['day']:
                    row['day'] = str(row['day'])
            
            return {
                "status": True,
                "message": "Cluster-wise daily trends fetched successfully",
                "data": data,
                "total_records": len(data)
            }
        except Exception as e:
            print("Exception in cluster_wise_daily_trends:", str(e))
            import traceback
            print("traceback:", traceback.format_exc())
            return {"status": False, "message": str(e), "data": [], "total_records": 0}

    @staticmethod
    async def risk_score_max_date(filters, cross_filters, drill_state, payload):
        try:
            resp = await urdhva_base.BasePostgresModel.get_aggr_data(
                query="SELECT MAX(version_date)::date AS max_date FROM public.tt_risk_score", skip_total=True
            )
            max_date = str(resp['data'][0]['max_date'])[:10] if resp.get('data') and resp['data'][0].get('max_date') else None
            return {"status": bool(max_date), "message": "Last updated date fetched successfully" if max_date else "No data found", "data": max_date}
        except Exception as e:
            return {"status": False, "message": str(e), "data": None}

    
    @staticmethod
    async def adding_device(filters, cross_filters, drill_state, payload):
        try:
            # Constants
            rpt = urdhva_base.context.context.get('rpt', {})
            sap_id = rpt.get("sap_id")
            print(f'sap_id: {sap_id}')
            sap_id = sap_id[0] 
            
            BLOCKED_CODES = {"0000010001", "00"}
            COLUMN_MAPPING = {
                "TRUCK_REGNNO": "SAP TT No.",
                "LOCN_CODE": "sap_id",
                "TRANS_ID": "Transporter",
                "bu": "Select Business",
                "name": "Location"
            }
            
            sap_tt_val = payload.get("sap_tt_no")
            safe_val = str(sap_tt_val).replace("'", "''")

            # Fetch truck records
            truck_query = f"""
                SELECT TRUCK_REGNNO, LOCN_CODE, TRANS_ID
                FROM "IMS_SAP"."TRUCK_DETAILS"
                WHERE TRUCK_REGNNO = '{safe_val}' and LOCN_CODE = '{sap_id}'
            """

            charts_ins = dashboard_studio_model.Charts_Connection_Vault_RoutingParams(
                connection_id=connection_mapping.connection_mapping.get("ims", "1"),
                action="execute_query"
            )

            function = await charts_actions.charts_connection_vault_routing(charts_ins)
            truck_result = await function(query=truck_query)

            truck_df = pl.DataFrame(truck_result, schema={
                "TRUCK_REGNNO": pl.String,
                "LOCN_CODE": pl.String,
                "TRANS_ID": pl.String,
            })

            if truck_df.height == 0:
                                
                # connection for vts_truck
                dashboard_studio_model.Charts_Connection_Vault_RoutingParams.connection_id = connection_mapping.connection_mapping.get("vts", "5")
                dashboard_studio_model.Charts_Connection_Vault_RoutingParams.action = 'execute_query'
                function = await charts_actions.charts_connection_vault_routing(dashboard_studio_model.Charts_Connection_Vault_RoutingParams)

                # check from completed_trip table
                query = f"""
                        SELECT top 1
                        vehicle_rto_no, depot_erp_code, erp_transporter_code
                        FROM completed_trip 
                        WHERE vehicle_rto_no = '{safe_val}' AND depot_erp_code = '{sap_id}'
                    """
                
                trip_completed_resp = await function(query=query)
                completed_trips = pl.DataFrame(trip_completed_resp)
                
                # Map the returned columns to the correct schema
                truck_df = completed_trips.rename({
                    "vehicle_rto_no": "TRUCK_REGNNO",
                    "depot_erp_code": "LOCN_CODE", 
                    "erp_transporter_code": "TRANS_ID"
                })
                              
                if truck_df.height == 0:
                    return {
                        "status": False,
                        "message": "No truck details found for given SAP TT No",
                        "data": {}
                    }
                
            # Filter blocked transporters
            truck_df = truck_df.filter(
                ~pl.col("TRANS_ID").cast(pl.String).str.strip_chars().is_in(BLOCKED_CODES)
            )

            if truck_df.height == 0:
                return {
                    "status": False,
                    "message": "trucks is blocked",
                    "data": {}
                }

            # Get unique SAP IDs for bulk fetch
            sap_ids = truck_df.select("LOCN_CODE").unique().to_series().to_list()
            sap_ids_string = ",".join(f"'{x}'" for x in sap_ids)

            # Fetch locations in bulk
            location_query = f"""
                SELECT sap_id,bu,name,zone
                FROM LOCATION_MASTER
                WHERE sap_id IN ({sap_ids_string})
            """
            location_result = await VTSAnalyticsActions.execute_query(location_query)
            location_df = pl.DataFrame(location_result, schema={
                "sap_id": pl.String,
                "bu": pl.String,
                "name": pl.String,
                "zone":pl.String
            })

            if location_df.height == 0:
                return {
                    "status": False,
                    "message": "No matching locations found",
                    "data": {}
                }

            # Join and rename
            final_df = truck_df.join(location_df, left_on="LOCN_CODE", right_on="sap_id", how="inner")
            
            if final_df.height == 0:
                return {
                    "status": False,
                    "message": "No valid truck-location mapping found","data": {}
                }
                
            # Execute Tibco query
                              
            shortage_tibco_query = f"""select ENGINE_NO, CHASSIS_NO, vehicle_no
            from veh_blklis_stg
            where vehicle_no = '{safe_val}'  """
            print('shortage_tibco_query------->',shortage_tibco_query)
            shortage_data = await VTSAnalyticsActions.execute_query(shortage_tibco_query)            
            print('shortage_data',shortage_data)
            
            # Merge shortage_data with final_df on TRUCK_REGNNO (final_df) = vehicle_no (shortage_data)
            if not shortage_data.empty:
                shortage_data_pl = pl.from_pandas(shortage_data)
                final_df = final_df.join(shortage_data_pl, left_on='TRUCK_REGNNO', right_on='vehicle_no', how='left')

            final_df = final_df.rename(COLUMN_MAPPING)
            # print('fina_df',final_df)

            return {
                "status": True, "message": "Data fetched successfully",
                "data": final_df.to_dicts()
            }

        except Exception as e:
            print("Exception:", str(e))
            return {
                "status": False,
                "message": f"Failed to fetch SAP truck details: {str(e)}",
                "data": {}
            }
    
    @staticmethod
    async def device_commissioning_table(filters, cross_filters, drill_state, payload):
        try:
            query = """select * from device_installation"""

            if not query:
                return {"status": False, "message": "Query not found", "data": []}

            # Execute VTS history query
            df = await VTSAnalyticsActions.execute_query(query)
            # print(df)

            return{"status" :True , "message":"success","data":df.to_dict(orient="records")}
        except Exception as e:
            return {"status": False, "message": str(e), "data": {}}
    
    @staticmethod
    async def vts_accept_and_block(filters, cross_filters, drill_state, payload):
        try:
            base_query = vts_query.vts_query.get("accept_and_block")
            condition = VTSAnalyticsActions.build_filter_conditions(filters, cross_filters, base_query)
            
            if isinstance(condition, list):
                condition = " AND ".join(condition)

            final_condition = ""
            if condition:
                condition = (
                    condition.replace("bu", "a.bu").replace("created_at", "a.created_at"))

                final_condition = " AND " + condition

            final_query = base_query.format(final_condition=final_condition)
            print("Final Query: ", final_query)

            merged_df = await VTSAnalyticsActions.execute_query(final_query, engine="polars")
            merged_df = (merged_df.explode("notices").unnest("notices"))
            
            # Unique alert_id sets
            system_ids = (
                merged_df.filter(pl.col("doc_type") == "System Generated").select("alert_id").unique())
                        
            user_ids = (
                merged_df.filter(pl.col("doc_type") == "User Created").select("alert_id").unique())

            # Compare alert_id sets
            system_only_ids = system_ids.join( user_ids, on = "alert_id", how = "anti")                                 
            system_only_df = (merged_df.join(system_only_ids, on = "alert_id", how = "inner").unique(subset=["alert_id"]))
            
            both_ids = system_ids.join( user_ids, on="alert_id", how="inner")
            both_df = ( merged_df.join(both_ids, on="alert_id", how="inner").unique(subset=["alert_id"]))
            
            system_only_df, both_df = (
                    system_only_df.drop("file_path","report_type", strict=False) , both_df.drop("file_path","report_type", strict=False))
            
            if payload.get("download") == "true":
                s = system_only_df.with_columns(pl.lit("no").alias("Show_Cause_Notice"))
                b = both_df.with_columns(pl.lit("yes").alias("Show_Cause_Notice"))
                combined = pl.concat([s, b], how="vertical")
                
                return await download_streaming_data(combined, filename='Show_Cause_Notice')
                
                        
            return {"status": True, "message": "success","data":{ "system_only" :system_only_df.height,"system_and_user":both_ids.height}}
        except Exception as e:
            print("Exception in vts_accept_and_block:", str(e))
            print("traceback:", traceback.format_exc())
            return {"status": False,"message": str(e),"data": []}

