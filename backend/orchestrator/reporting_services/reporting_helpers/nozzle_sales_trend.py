
import urdhva_base
import polars as pl
import numpy as np
import os
import datetime
import utilities.connection_mapping as connection_mapping
from charts_actions import charts_connection_vault_routing
from dashboard_studio_model import Charts_Connection_Vault_RoutingParams
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.patches as patches
from matplotlib.font_manager import FontProperties


async def plot_ms_sales_trend(trend_data: pl.DataFrame, output_path="/tmp/nozzle_trend_chart.png"):

    if trend_data is None or trend_data.is_empty():
        raise ValueError("trend_data cannot be empty")

    # -----------------------------
    # PREP DATA
    # -----------------------------

    df = trend_data.with_columns([
        pl.col("transaction_date").dt.strftime("%d-%b").alias("day"),
        ((pl.col("ms_power").cast(pl.Float64) /
         pl.col("ms_total").cast(pl.Float64))
        * 100
        ).alias("conversion")
        ])

    # Remove nulls
    df = df.filter(
        pl.col("ms_total").is_not_null() &
        pl.col("ms_power").is_not_null()
    )

    # Convert Decimal → Float
    df = df.with_columns([
        pl.col("ms_total").cast(pl.Float64),
        pl.col("ms_power").cast(pl.Float64),
        pl.col("conversion").cast(pl.Float64)
    ])

    # Clamp conversion to avoid extreme values
    df = df.with_columns([
        pl.when(pl.col("conversion") > 100)
        .then(100)
        .otherwise(pl.col("conversion"))
        .alias("conversion")
    ])

    # -----------------------------
    # NUMPY ARRAYS (SAFE)
    # -----------------------------
    days = df["day"].to_numpy()

    ms_total = np.nan_to_num(df["ms_total"].to_numpy(), nan=0.0, posinf=0.0, neginf=0.0)
    ms_power = np.nan_to_num(df["ms_power"].to_numpy(), nan=0.0, posinf=0.0, neginf=0.0)
    conversion = np.nan_to_num(df["conversion"].to_numpy(), nan=0.0, posinf=0.0, neginf=0.0)

    x = np.arange(len(days))

    # -----------------------------
    # AVERAGES
    # -----------------------------
    avg_total = float(np.mean(ms_total)) if len(ms_total) > 0 else 0
    avg_conversion = float(np.mean(conversion)) if len(conversion) > 0 else 0

    # -----------------------------
    # FIGURE (SAFE SIZE)
    # -----------------------------
    fig, ax1 = plt.subplots(figsize=(14, 5))

    # -----------------------------
    # BARS
    # -----------------------------
    ax1.bar(x - 0.2, ms_total, width=0.4, color="#5b7db1", label="MS Total")
    ax1.bar(x + 0.2, ms_power, width=0.4, color="#c9c9c9", label="MS Power")

    # -----------------------------
    # LINE
    # -----------------------------
    ax2 = ax1.twinx()
    ax2.plot(x, conversion, marker='o',color="#d4aa00", linewidth=2, label="MS % Conversion")

    # -----------------------------
    # AXES
    # -----------------------------
    ax1.set_xticks(x)
    #ax1.set_xticklabels(days)
    ax1.set_xticklabels(days, rotation=45)

    ax2.set_ylim(0, 25)
    ticks = np.arange(0, 26, 5)
    ax1.set_ylabel("Sales (TMT)")
    ax2.set_ylabel("Conversion (%)")
    ax2.set_yticks(ticks)
    ax2.set_yticklabels([f"{int(t)}%" for t in ticks])

    # -----------------------------
    # AVERAGE LINES
    # -----------------------------
    ax1.axhline(avg_total, linestyle="--", color="#bfbfbf")
    ax2.axhline(avg_conversion, linestyle=":", color="#d4aa00")

    # -----------------------------
    # LABELS
    # -----------------------------
    offset = max(ms_total) * 0.02 if len(ms_total) > 0 else 1

    for i, v in enumerate(ms_total):
        ax1.text(i - 0.2, v + offset, f"{int(v)}", fontsize=7, ha='center')

    for i, v in enumerate(ms_power):
        ax1.text(i + 0.2, v + offset, f"{v:.1f}", fontsize=7, ha='center')

    for i, v in enumerate(conversion):
        #if i % 2 == 0:
        ax2.text(i, float(v) + 0.3, f"{float(v):.1f}%", fontsize=9, ha='center', fontweight='bold')

    # -----------------------------
    # STYLE
    # -----------------------------
    ax1.grid(axis='y', linestyle='--', alpha=0.3)

    for spine in ["top", "right"]:
        ax1.spines[spine].set_visible(False)
        ax2.spines[spine].set_visible(False)

    start_date = df["transaction_date"].min()
    end_date = df["transaction_date"].max()

    start_label = start_date.strftime("%d-%b")
    end_label = end_date.strftime("%d-%b")
    plt.title(f"MS Total vs Power with Conversion % (TMT) ({start_label} to {end_label})", pad=25)

    fig.legend(loc="upper left", bbox_to_anchor=(0.01, 0.87), ncol=3, frameon=False)
    plt.tight_layout(rect=[0, 0, 1, 0.92])

    # -----------------------------
    # SAVE
    # -----------------------------
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()

    return output_path


def wrap_zone_name(name: str) -> str:
    """Split zone names with a space onto two lines (e.g. 'Bhubaneswar Zone' -> 'Bhubaneswar\nZone').
    Names with no space (EZ, NCZ, NWZ, NZ, SCZ, SZ, WZ) stay on one line."""
    if " " in name:
        parts = name.rsplit(" ", 1)
        return f"{parts[0]}\n{parts[1]}"
    return name


async def conversion_graph(data: pl.DataFrame, output_path=None, type = str):
    output_path = f"/tmp/nozzle_{type}_conv.png"
    if data is None or data.is_empty():
        raise ValueError("trend_data cannot be empty")

    data = data.with_columns(
        pl.col("month").str.strptime(pl.Date, "%b'%y").alias("month_date")
    )

    month_df = (
        data.select(["month", "month_date"])
        .unique()
        .sort("month_date")
    )
    months = month_df.sort("month_date", descending=True)["month"].to_list()

    if type == "power" or type == "power_6":
        try:
            pivot_df = (
                data.select(["month", "zone", "% POWER Conversion"])
                .pivot(values="% POWER Conversion", index="zone", on="month")
                .fill_null(0)
            )
        except TypeError:
            pivot_df = (
                data.select(["month", "zone", "% POWER Conversion"])
                .pivot(values="% POWER Conversion", index="zone", columns="month")
                .fill_null(0)
            )
    elif type == "turbo":
        try:
            pivot_df = (
                data.select(["month", "zone", "% TURBO Conversion"])
                .pivot(values="% TURBO Conversion", index="zone", on="month")
                .fill_null(0)
            )
        except TypeError:
            pivot_df = (
                data.select(["month", "zone", "% TURBO Conversion"])
                .pivot(values="% TURBO Conversion", index="zone", columns="month")
                .fill_null(0)
            )
            
    pivot_df = pivot_df.with_columns([
        pl.col(m).cast(pl.Float64) for m in pivot_df.columns if m != "zone"
    ])

    pivot_df = pivot_df.with_columns(
        pl.col("zone").str.to_uppercase().eq("PAN INDIA").alias("_is_pan_india")
    ).sort(["_is_pan_india", "zone"]).drop("_is_pan_india")

    zones = pivot_df["zone"].to_list()
    values = [pivot_df[m].to_list() for m in months]

    LABEL_COL_WIDTH = 0.05

    x = np.arange(len(zones)) * 1.2
    group_width = 0.8
    bar_width = group_width / len(months)

    fig, ax = plt.subplots(figsize=(45, 15))

    colors = ["#4472C4", "#ED7D31", "#A5A5A5", "#FFC000", "#5B9BD5", "#70AD47"]
    
    
    for i, month in enumerate(months):
        ax.bar(
            x - group_width / 2 + i * bar_width + bar_width / 2,
            values[i], width=bar_width, color=colors[i % len(colors)], label=month
        )

    tick_pos = x + bar_width * (len(months) - 1) / 2
    ax.set_xticks(tick_pos)
    ax.set_xticklabels([])
    ax.tick_params(axis="x", length=0) # tick lines disappear
    ax.tick_params(axis="y", labelsize=18)
    ax.set_ylabel("Conversion %", fontsize=20, weight="bold")
    if type == "power" or type == "power_6":
        ax.set_title("% POWER Conversion", fontsize=24, weight="bold")
    elif type == "turbo":
        ax.set_title("% TURBO Conversion", fontsize=24, weight="bold")
    ax.yaxis.set_major_formatter(lambda x, _: f"{x:.0f}%")
    ax.set_xlim(x.min() - 0.5, x.max() + 0.5)
    max_val = max(max(v) for v in values)
    ax.set_ylim(0, max_val * 1.15)

    col_labels = [wrap_zone_name(z) for z in zones]
    table_data = []
    for m in months:
        row = [f"{v:.2f}%" for v in pivot_df[m].to_list()]
        table_data.append(row)

    header_row_h = 0.2
    data_row_h = 0.08
    table_bbox_height = header_row_h + data_row_h * len(months)

    table = ax.table(
        cellText=table_data, rowLabels=months, colLabels=col_labels,
        cellLoc="center", loc="bottom", bbox=[0, -table_bbox_height, 1, table_bbox_height]
    )
    table.auto_set_font_size(False)
    table.set_fontsize(20)

    month_col_width = LABEL_COL_WIDTH
    zone_col_width = (1 - month_col_width) / len(zones)

    for (row, col), cell in table.get_celld().items():
        cell.set_text_props(rotation=0, ha="center", va="center")
        if row == 0:
            cell.set_height(header_row_h / table_bbox_height)
        else:
            cell.set_height(data_row_h / table_bbox_height)
        if col == -1:
            cell.set_width(month_col_width)
            if row >= 1:
                cell.set_text_props(ha="right", va="center", fontweight="bold")
                cell.PAD = 0.3
        else:
            cell.set_width(zone_col_width)

    top = 0.95
    footer_gap = 0.04
    margin = 0.02
    bottom = (table_bbox_height * top + footer_gap + margin) / (1 + table_bbox_height)
    axes_left, axes_right = 0.2, 0.98
    plt.subplots_adjust(left=axes_left, right=axes_right, top=top, bottom=bottom)

    # -----------------------------
    # Color swatches — a small square next to each month's row label.
    #
    # Position: matplotlib renders the row-label column (col=-1) to the
    # LEFT of the table's nominal x=0 — it does NOT start at axes_left.
    # So we read back the cell's *actual* rendered position via
    # get_window_extent (ground truth) rather than assuming a location.
    #
    # Size: computed with separate x/y scaling based on real figure
    # width vs height, so the swatch renders as a true square (not a
    # smeared bar) even on a very wide, short figure like this one.
    # -----------------------------
    axes_height_fig = top - bottom
    row_height_fig_frac = data_row_h * axes_height_fig

    fig_w_in, fig_h_in = fig.get_size_inches()
    row_height_inches = row_height_fig_frac * fig_h_in

    fig.canvas.draw()
    renderer = fig.canvas.get_renderer()

    # Measure the widest month label at the actual row-label font (20, bold)
    label_font = FontProperties(size=20, weight="bold")
    max_text_w_px = max(
        renderer.get_text_width_height_descent(m, label_font, ismath=False)[0]
        for m in months
    )
    text_w_inches = max_text_w_px / fig.dpi

    side_inches = row_height_inches * 0.45
    gap_inches = row_height_inches * 0.3
    margin_inches = row_height_inches * 0.35  # inner breathing room on both sides

    needed_width_inches = text_w_inches + gap_inches + side_inches + 2 * margin_inches
    axes_width_inches = (axes_right - axes_left) * fig_w_in
    needed_col_frac = needed_width_inches / axes_width_inches

    month_col_width = max(LABEL_COL_WIDTH, needed_col_frac)
    zone_col_width = (1 - month_col_width) / len(zones)
    for i in range(len(months)):
        cell = table[(i + 1, -1)]
        cell_bbox_fig = cell.get_window_extent(renderer).transformed(fig.transFigure.inverted())

        text_artist = cell.get_text()
        text_w_px, _, _ = renderer.get_text_width_height_descent(
            text_artist.get_text(), text_artist.get_fontproperties(), ismath=False,
        )
        text_w_frac = (text_w_px / fig.dpi) / fig_w_in

        cell_center = cell_bbox_fig.x0 + cell_bbox_fig.width / 2
        text_right = cell_center + text_w_frac / 2

        side_x_frac = side_inches / fig_w_in
        side_y_frac = side_inches / fig_h_in
        gap_x_frac = gap_inches / fig_w_in

        sq_x = text_right + gap_x_frac
        # Hard clamp: swatch's right edge can never cross the cell's right edge
        max_sq_x = cell_bbox_fig.x1 - (margin_inches / fig_w_in) - side_x_frac
        sq_x = min(sq_x, max_sq_x)

        sq_y = cell_bbox_fig.y0 + (cell_bbox_fig.height - side_y_frac) / 2

        fig.add_artist(patches.Rectangle(
            (sq_x, sq_y), side_x_frac, side_y_frac,
            transform=fig.transFigure,
            facecolor=colors[i % len(colors)],
            edgecolor="none",
            zorder=10,
        ))

    latest_row = (
        data.select(["month", "month_date", "months_with_dates"])
        .unique(subset=["month"])
        .sort("month_date")
        .row(-1)
    )
    latest_month, latest_period = latest_row[0], latest_row[2]
    footer_text = f"{latest_month} Period : {latest_period}"
    plt.figtext(0.98, margin, footer_text, ha="right", fontsize=20, weight="bold")

    plt.savefig(output_path, bbox_inches="tight")
    plt.close()
    print("output_path: ", output_path)
    return output_path


def create_nozzle_excel(nozzle_sales_avg_excel: list, periods: list, output_path="/tmp/nozzle_sales_categorized.xlsx"):

    wb = Workbook()
    ws = wb.active
    ws.title = "Nozzle Sales"

    data_map = {row["period"].strip(): row for row in nozzle_sales_avg_excel}

    header_fill = PatternFill(start_color="C9CFDB", end_color="C9CFDB", fill_type="solid")
    category_fill = PatternFill(start_color="DBE5D6", end_color="DBE5D6", fill_type="solid")
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Category"] + periods

    ws.cell(row=1, column=1, value="Category")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    ws.cell(row=1, column=2, value="Nozzle Sales in MT")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(headers))

    for col, header in enumerate(periods, start=2):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).border = border
        ws.cell(row=2, column=col).border = border

    cat_cell = ws.cell(row=1, column=1)
    cat_cell.font = bold_font
    cat_cell.alignment = center_align

    top_cell = ws.cell(row=1, column=2)
    top_cell.font = bold_font
    top_cell.alignment = center_align

    rows_structure = [
        ("MS-R Normal", "MS-R Normal"),
        ("MS-R Branded", "MS-R Branded"),
        ("% Conversion", "MS %"),
        ("MS-Total", "MS Total"),

        ("HSD-R Normal", "HSD-R Normal"),
        ("HSD-R Branded", "HSD-R Branded"),
        ("% Conversion", "HSD %"),
        ("HSD-Total", "HSD Total"),
    ]

    row_idx = 3

    for label, key_prefix in rows_structure:
        ws.cell(row=row_idx, column=1, value=label)

        for col_idx, period in enumerate(periods, start=2):

            row = data_map.get(period.strip(), {})
            value = ""

            if row:
                if key_prefix == "MS-R Normal":
                    value = row.get("MS-R Normal", "")
                elif key_prefix == "MS-R Branded":
                    value = row.get("MS-R Branded", "")
                elif key_prefix == "MS %":
                    value = row.get("% MS Conversion", "")
                elif key_prefix == "MS Total":
                    value = row.get("MS-Total", "")

                elif key_prefix == "HSD-R Normal":
                    value = row.get("HSD-R Normal", "")
                elif key_prefix == "HSD-R Branded":
                    value = row.get("HSD-R Branded", "")
                elif key_prefix == "HSD %":
                    value = row.get("% HSD Conversion", "")
                elif key_prefix == "HSD Total":
                    value = row.get("HSD-Total", "")

            # Add % sign
            if "%" in key_prefix and value != "":
                value = f"{value}%"

            ws.cell(row=row_idx, column=col_idx, value=value)


        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            if col_idx == 1:
                cell.fill = category_fill
                cell.font = bold_font
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if "Total" in label:
                cell.fill = total_fill
                cell.font = bold_font

        row_idx += 1

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)

        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 4

    wb.save(output_path)
    return output_path

    
async def fetch_data():

    nozzle_sales_query_avg = f"""                 
                            WITH base AS (
                                SELECT
                                    "transaction_date",
                                    SUM("sales_volume") FILTER (WHERE product_grp = 'MS') AS ms,
                                    SUM("sales_volume") FILTER (WHERE product_grp IN ('POWER 99','POWER 95','POWER 100')) AS power,
                                    SUM("sales_volume") FILTER (WHERE product_grp = 'HSD') AS hsd,
                                    SUM("sales_volume") FILTER (WHERE product_grp = 'TURBO') AS turbo
                                FROM public.nozzle_sales
                                WHERE "transaction_date" >= DATE_TRUNC('year', CURRENT_DATE) + INTERVAL '2 months'
                                AND "transaction_date"::DATE < CURRENT_DATE
                                AND sap_id LIKE '4%' AND zone is not NULL
                                GROUP BY "transaction_date"
                            ),

                            final_data AS (
                                SELECT
                                    DATE_TRUNC('month', "transaction_date") AS month_start,
                                    TO_CHAR(DATE_TRUNC('month', "transaction_date"), 'FMDDth Mon')
                                    || ' to '
                                    || TO_CHAR(
                                        LEAST(
                                            (DATE_TRUNC('month', "transaction_date") + INTERVAL '1 month - 1 day')::DATE,
                                            CURRENT_DATE - INTERVAL '1 day'
                                        ),
                                        'FMDDth Mon'
                                    ) 
                                    || ' Avg ' AS period,
                                    ROUND(((AVG(ms))    / 1411.0) / 0.89)::BIGINT AS ms_normal,
                                    ROUND(((AVG(power)) / 1411.0) / 0.89)::BIGINT AS ms_branded,
                                    ROUND(((AVG(hsd))   / 1210.0) / 0.89)::BIGINT AS hsd_normal,
                                    ROUND(((AVG(turbo)) / 1210.0) / 0.89)::BIGINT AS hsd_branded,

                                    1 AS sort_order

                                FROM base
                                GROUP BY DATE_TRUNC('month', "transaction_date")

                                UNION ALL

                                SELECT
                                    CURRENT_DATE - INTERVAL '1 day' AS month_start,
                                    TO_CHAR(CURRENT_DATE - INTERVAL '1 day', 'FMDDth Mon'),

                                    ROUND(((AVG(ms))    / 1411.0) / 0.89)::BIGINT,
                                    ROUND(((AVG(power)) / 1411.0) / 0.89)::BIGINT,
                                    ROUND(((AVG(hsd))   / 1210.0) / 0.89)::BIGINT,
                                    ROUND(((AVG(turbo)) / 1210.0) / 0.89)::BIGINT,
                                    2
                                FROM base
                                WHERE "transaction_date"::DATE = CURRENT_DATE - INTERVAL '1 day'
                            )

                            SELECT
                                period,

                                ms_normal     AS "MS-R Normal",
                                ms_branded    AS "MS-R Branded",

                                ROUND(
                                    (ms_branded * 100.0) / NULLIF(ms_normal + ms_branded, 0),
                                    2
                                ) AS "% MS Conversion",

                                (ms_normal + ms_branded) AS "MS-Total",

                                hsd_normal    AS "HSD-R Normal",
                                hsd_branded   AS "HSD-R Branded",

                                ROUND(
                                    (hsd_branded * 100.0) / NULLIF(hsd_normal + hsd_branded, 0),
                                    2
                                ) AS "% HSD Conversion",

                                (hsd_normal + hsd_branded) AS "HSD-Total"

                            FROM final_data
                            ORDER BY month_start;
                        """

    nozzle_sales_conversion_query = f""" 
                WITH base AS (
                    SELECT
                        transaction_date, zone,
                        SUM(sales_volume) FILTER (WHERE product_grp IN ('MS')) AS ms,
                        SUM(sales_volume) FILTER (WHERE product_grp IN ('POWER 99','POWER 95','POWER 100')) AS power,
                        SUM(sales_volume) FILTER (WHERE product_grp IN ('HSD')) AS hsd,
                        SUM(sales_volume) FILTER (WHERE product_grp IN ('TURBO')) AS turbo
                    FROM public.nozzle_sales
                    WHERE 
                    transaction_date::DATE >= 
                        CASE 
                            WHEN CURRENT_DATE = DATE_TRUNC('year', CURRENT_DATE)::DATE
                            THEN DATE_TRUNC('year', CURRENT_DATE - INTERVAL '1 year')  -- 2026-01-01
                            ELSE DATE_TRUNC('year', CURRENT_DATE)                      -- 2027-01-01
                        END
                    AND transaction_date::DATE < 
                        CASE 
                            WHEN CURRENT_DATE = DATE_TRUNC('year', CURRENT_DATE)::DATE
                            THEN DATE_TRUNC('year', CURRENT_DATE)                      -- 2027-01-01
                            ELSE CURRENT_DATE::DATE                                          -- today
                        END
                    AND sap_id LIKE '4%' AND zone IS NOT NULL
                    GROUP BY transaction_date, zone
                ),
                final_data AS (
                    SELECT
                        DATE_TRUNC('month', transaction_date) AS month_date, zone,
                        TO_CHAR(DATE_TRUNC('month', transaction_date), 'FMDDth Mon')
                        || ' to '
                        || TO_CHAR(
                            LEAST(
                            (DATE_TRUNC('month', "transaction_date") + INTERVAL '1 month - 1 day')::DATE,
                            CURRENT_DATE - INTERVAL '1 day'
                            ),
                            'FMDDth Mon'
                        ) AS months_with_dates,
                        TO_CHAR(DATE_TRUNC('month', transaction_date), 'Mon''YY') AS month,
                        ROUND(((AVG(ms)) / 1411.0) / 0.89)::BIGINT AS ms_normal,
                        ROUND(((AVG(power)) / 1411.0) / 0.89)::BIGINT AS ms_branded,
                        ROUND(((AVG(hsd)) / 1210.0) / 0.89)::BIGINT AS hsd_normal,
                        ROUND(((AVG(turbo)) / 1210.0) / 0.89)::BIGINT AS hsd_branded
                    FROM base
                    GROUP BY month_date, zone
                ),

                
                pan_india_base AS (
                    SELECT
                        transaction_date,
                        SUM(sales_volume) FILTER (WHERE product_grp IN ('MS')) AS ms,
                        SUM(sales_volume) FILTER (WHERE product_grp IN ('POWER 99','POWER 95','POWER 100')) AS power,
                        SUM(sales_volume) FILTER (WHERE product_grp IN ('HSD')) AS hsd,
                        SUM(sales_volume) FILTER (WHERE product_grp IN ('TURBO')) AS turbo
                    FROM public.nozzle_sales
                    WHERE 
                    transaction_date::DATE >= 
                        CASE 
                            WHEN CURRENT_DATE = DATE_TRUNC('year', CURRENT_DATE)::DATE
                            THEN DATE_TRUNC('year', CURRENT_DATE - INTERVAL '1 year')
                            ELSE DATE_TRUNC('year', CURRENT_DATE)
                        END
                    AND transaction_date::DATE < 
                        CASE 
                            WHEN CURRENT_DATE = DATE_TRUNC('year', CURRENT_DATE)::DATE
                            THEN DATE_TRUNC('year', CURRENT_DATE)
                            ELSE CURRENT_DATE::DATE
                        END
                    AND sap_id LIKE '4%'
                    GROUP BY transaction_date
                ),

                pan_india AS (
                    SELECT
                        TO_CHAR(DATE_TRUNC('month', transaction_date), 'Mon''YY') AS month,
                        TO_CHAR(DATE_TRUNC('month', transaction_date), 'FMDDth Mon')
                        || ' to '
                        || TO_CHAR(
                            LEAST(
                            (DATE_TRUNC('month', transaction_date) + INTERVAL '1 month - 1 day')::DATE,
                            CURRENT_DATE - INTERVAL '1 day'
                            ),
                            'FMDDth Mon'
                        ) AS months_with_dates,
                        'PAN INDIA' AS zone,
                        ROUND(((AVG(ms)) / 1411.0) / 0.89)::BIGINT AS "MS-R Normal",
                        ROUND(((AVG(power)) / 1411.0) / 0.89)::BIGINT AS "MS-R Branded",
                        ROUND(((AVG(hsd)) / 1210.0) / 0.89)::BIGINT AS "HSD-R Normal",
                        ROUND(((AVG(turbo)) / 1210.0) / 0.89)::BIGINT AS "HSD-R Branded"
                    FROM pan_india_base
                    GROUP BY DATE_TRUNC('month', transaction_date)
                )

                SELECT
                    month, months_with_dates, zone,
                    -- ms_normal AS "MS-R Normal",
                    -- ms_branded AS "MS-R Branded",
                    -- hsd_normal AS "HSD-R Normal",
                    -- hsd_branded AS "HSD-R Branded",
                    COALESCE(ROUND((ms_branded * 100.0) / NULLIF(ms_normal + ms_branded, 0), 2), 0) AS "% POWER Conversion",
                    COALESCE(ROUND((hsd_branded * 100.0) / NULLIF(hsd_normal + hsd_branded, 0), 2), 0) AS "% TURBO Conversion"
                FROM final_data
                
                UNION ALL

                SELECT
                    month, months_with_dates, zone,
                    -- "MS-R Normal",
                    -- "MS-R Branded", "HSD-R Normal", "HSD-R Branded",
                    COALESCE(
                        ROUND(
                            ("MS-R Branded" * 100.0 / NULLIF("MS-R Normal" + "MS-R Branded", 0))::numeric,
                            2
                        ),
                        0
                    ) AS "% POWER Conversion",
                    COALESCE(
                        ROUND(
                            ("HSD-R Branded" * 100.0 / NULLIF("HSD-R Normal" + "HSD-R Branded", 0))::numeric,
                            2
                        ),
                        0
                    ) AS "% TURBO Conversion"
                FROM pan_india

                ORDER BY zone
                """
                
    nozzle_trend_query = """
                SELECT
                    transaction_date,
                    ROUND(((SUM("sales_volume") FILTER (WHERE product_grp in ('MS','POWER 99','POWER 95','POWER 100'))/ 1411.0
                            ) / 1000.0
                        ) / 0.89, 2
                        ) AS ms_total,
                    ROUND(((SUM("sales_volume") FILTER (WHERE product_grp in ('POWER 99','POWER 95','POWER 100'))/ 1411.0
                            ) / 1000.0
                        ) / 0.89, 2
                        ) AS ms_power
                FROM public.nozzle_sales
                WHERE transaction_date >= CURRENT_DATE - INTERVAL '30 days'
                    AND transaction_date::DATE < CURRENT_DATE
                GROUP BY transaction_date

            """
    nozzle_sync_time_query = """ SELECT MAX(created_at::timestamp) as sync_time FROM nozzle_sales """
    Charts_Connection_Vault_RoutingParams.connection_id = connection_mapping.connection_mapping.get("hpcl_ceg", "1")
    Charts_Connection_Vault_RoutingParams.action = 'execute_query'
    function = await charts_connection_vault_routing(Charts_Connection_Vault_RoutingParams)


    nozzle_sales_avg = await function(query=nozzle_sales_query_avg)
    nozzle_sales_avg = pl.DataFrame(nozzle_sales_avg)
    
    nozzle_sales_avg_excel = nozzle_sales_avg.to_dicts()
    periods_excel = [row["period"] for row in nozzle_sales_avg_excel]
    
    today = datetime.date.today()
    months = []
    temp_date = today.replace(day=1)

    for _ in range(6):
        months.append(temp_date.strftime("%b")) 
        temp_date = (temp_date - datetime.timedelta(days=1)).replace(day=1)
    
    def format_date(d: datetime.date) -> str:
        day = d.day
        suffix = "th" if 11 <= day % 100 <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
        return f"{day}{suffix} {d.strftime('%b')}"

    yesterday_str = format_date(today - datetime.timedelta(days=1))
    # yesterday_str = (today - datetime.timedelta(days=1)).strftime("%-dth %b")
    print(yesterday_str)
    nozzle_sales_avg_df = nozzle_sales_avg.with_columns(
        pl.col("period").str.strip_chars().alias("period")
    ).filter(
        (
            (pl.col("period").str.contains("Avg")) &
            (pl.col("period").str.contains("|".join(months)))
        ) |
        (pl.col("period") == yesterday_str)
    )
    pivot = {}

    for row in nozzle_sales_avg_df.iter_rows(named=True):
        period = row["period"]

        pivot[f"MS-R Normal ({period})"] = row["MS-R Normal"] #"MS-R Normal (Apr)": 10
        pivot[f"MS-R Branded ({period})"] = row["MS-R Branded"]
        pivot[f"MS % ({period})"] = round(row["% MS Conversion"], 2)
        pivot[f"MS Total ({period})"] = row["MS-Total"]

        pivot[f"HSD-R Normal ({period})"] = row["HSD-R Normal"]
        pivot[f"HSD-R Branded ({period})"] = row["HSD-R Branded"]
        pivot[f"HSD % ({period})"] = round(row["% HSD Conversion"], 2)
        pivot[f"HSD Total ({period})"] = row["HSD-Total"]

    periods = list(nozzle_sales_avg_df["period"])
    nozzle_sales_avg_df = pivot
    nozzle_sales_excel = create_nozzle_excel(nozzle_sales_avg_excel, periods_excel)
    
    nozzle_trend = await function(query= nozzle_trend_query)
    nozzle_trend_df = pl.DataFrame(nozzle_trend)

    nozzle_trend_chart = await plot_ms_sales_trend(nozzle_trend_df)

    nozzle_sync_time = await function(query= nozzle_sync_time_query)
    nozzle_sync_time = pl.DataFrame(nozzle_sync_time)
    nozzle_sync_time = nozzle_sync_time.with_columns(
        pl.col("sync_time")
        .dt.replace_time_zone("UTC")
        .dt.convert_time_zone("Asia/Kolkata")
        .dt.strftime("%-I:%M %p")
        .alias("nozzle_sales_sync")
    )
    print("nozzle_sync_time -------------->\n", nozzle_sync_time.to_dicts())
    
    nozzle_sales_conversion = await function(query=nozzle_sales_conversion_query)
    nozzle_sales_conversion_df = pl.DataFrame(nozzle_sales_conversion)
    
    #
    def month_limitation(df, months_limit=6):
        return (
            df
            .with_columns(
                pl.col("month").str.strptime(pl.Date, "%b'%y").alias("month_date")
            )
            .sort(["zone", "month_date"])
            .group_by("zone")
            .tail(months_limit)
        )
        
    graph_3_months = month_limitation(nozzle_sales_conversion_df, months_limit=3) # last 3 months
    graph_6_months = month_limitation(nozzle_sales_conversion_df, months_limit=6) # last 6 months

    nozzle_sales_power_conversion_df = graph_3_months.select(["zone", "month", "% POWER Conversion", "months_with_dates"])
    nozzle_sales_turbo_conversion_df = graph_3_months.select(["zone", "month", "% TURBO Conversion", "months_with_dates"])
    nozzle_sales_power_conversion_6 = graph_6_months.select(["zone", "month", "% POWER Conversion", "months_with_dates"])
    
    nozzle_power_conversion_graph = await conversion_graph(nozzle_sales_power_conversion_df, type="power")
    nozzle_turbo_conversion_graph = await conversion_graph(nozzle_sales_turbo_conversion_df, type="turbo")
    nozzle_power_conversion_graph_6 = await conversion_graph(nozzle_sales_power_conversion_6, type="power_6") 
    
    power_pivot_df = (
            nozzle_sales_power_conversion_6
            .with_columns([
                pl.col("% POWER Conversion").cast(pl.Float64)
            ])
            .pivot(
                values="% POWER Conversion",
                index="month",
                columns="zone"
            )
        )
    
    power_pivot_df = power_pivot_df.with_columns([
        pl.col("month").map_elements(
            lambda x: datetime.datetime.strptime(x, "%b'%y")
        ).alias("month_sort")
    ]).sort("month_sort", descending=True).drop("month_sort")
    power_pivot_df = power_pivot_df.head(6)
    cols = power_pivot_df.columns
    zone_cols = [c for c in cols if c not in ["month", "PAN INDIA"]]
    zone_cols_sorted = sorted(zone_cols)

    final_cols = ["month"] + zone_cols_sorted + ["PAN INDIA"]

    power_pivot_df = power_pivot_df.select(final_cols)
    power_pivot_df = power_pivot_df.rename({"month": "Month"})

    power_data = power_pivot_df.to_dicts()
    power_data = [
        {
            k: f"{v:.2f}%" if isinstance(v, float) else v
            for k, v in row.items()
        }
        for row in power_pivot_df.to_dicts()
    ]
    power_columns = power_pivot_df.columns
    
    return {"nozzle_sales_avg_df": nozzle_sales_avg_df, "periods" :periods, "nozzle_trend_chart": nozzle_trend_chart,
            "nozzle_sales_sync_time": nozzle_sync_time["nozzle_sales_sync"][0], "power_conversion_graph_3": nozzle_power_conversion_graph,
            "nozzle_sales_excel": nozzle_sales_excel, "power_data": power_data, "power_columns": power_columns, 
            "turbo_conversion_graph_3": nozzle_turbo_conversion_graph, "power_conversion_graph": nozzle_power_conversion_graph_6
        }


async def nozzles_sales_top_performance():

    nozzle_sales_top_query = f""" 
                                SELECT
                                    zone,
                                    region,
                                    sales_area,
                                    sap_id,
                                    location_name,

                                    /* ================= CURRENT MONTH ================= */

                                    ROUND(
                                        SUM(sales_volume) FILTER (
                                            WHERE product_grp IN ('POWER 99', 'POWER 95', 'POWER 100')
                                            AND transaction_date >= 
                                                CASE
                                                    WHEN CURRENT_DATE = date_trunc('month', CURRENT_DATE)::DATE
                                                    THEN date_trunc('month', CURRENT_DATE - INTERVAL '1 month')
                                                    ELSE date_trunc('month', CURRENT_DATE)
                                                END
                                            AND transaction_date <
                                                CASE
                                                    WHEN CURRENT_DATE = date_trunc('month', CURRENT_DATE)::DATE
                                                    THEN date_trunc('month', CURRENT_DATE)
                                                    ELSE CURRENT_DATE::DATE
                                                END
                                        ) / 1000.0,
                                        2
                                    ) AS current_power_kl,

                                    ROUND(
                                        SUM(sales_volume) FILTER (
                                            WHERE product_grp IN ('MS')
                                            AND transaction_date >= 
                                                CASE
                                                    WHEN CURRENT_DATE = date_trunc('month', CURRENT_DATE)::DATE
                                                    THEN date_trunc('month', CURRENT_DATE - INTERVAL '1 month')
                                                    ELSE date_trunc('month', CURRENT_DATE)
                                                END
                                            AND transaction_date <
                                                CASE
                                                    WHEN CURRENT_DATE = date_trunc('month', CURRENT_DATE)::DATE
                                                    THEN date_trunc('month', CURRENT_DATE)
                                                    ELSE CURRENT_DATE::DATE
                                                END
                                        ) / 1000.0,
                                        2
                                    ) AS current_ms_kl,

                                    /* ================= YTD ================= */

                                    ROUND(
                                        SUM(sales_volume) FILTER (
                                            WHERE product_grp IN ('POWER 99', 'POWER 95', 'POWER 100')
                                            AND transaction_date >= 
                                                CASE
                                                    WHEN EXTRACT(MONTH FROM CURRENT_DATE) >= 4
                                                    THEN date_trunc('year', CURRENT_DATE) + INTERVAL '3 month'
                                                    ELSE date_trunc('year', CURRENT_DATE - INTERVAL '1 year') + INTERVAL '3 month'
                                                END
                                            AND transaction_date::DATE < CURRENT_DATE
                                        ) / 1000.0,
                                        2
                                    ) AS ytd_power_kl,

                                    ROUND(
                                        SUM(sales_volume) FILTER (
                                            WHERE product_grp = 'MS'
                                            AND transaction_date >= 
                                                CASE
                                                    WHEN EXTRACT(MONTH FROM CURRENT_DATE) >= 4
                                                    THEN date_trunc('year', CURRENT_DATE) + INTERVAL '3 month'
                                                    ELSE date_trunc('year', CURRENT_DATE - INTERVAL '1 year') + INTERVAL '3 month'
                                                END
                                            AND transaction_date::DATE < CURRENT_DATE
                                        ) / 1000.0,
                                        2
                                    ) AS ytd_ms_kl,

                                    CASE
                                        WHEN CURRENT_DATE = date_trunc('month', CURRENT_DATE)::DATE THEN
                                            TO_CHAR(date_trunc('month', CURRENT_DATE - INTERVAL '1 month'), 'Mon''YY')
                                        ELSE
                                            TO_CHAR(date_trunc('month', CURRENT_DATE), 'FMDDth') || ' to ' ||
                                            TO_CHAR(CURRENT_DATE - INTERVAL '1 day', 'FMDDth Mon''YY')
                                    END AS current_date_label,

                                    TO_CHAR(
                                        CASE
                                            WHEN EXTRACT(MONTH FROM CURRENT_DATE) >= 4
                                            THEN date_trunc('year', CURRENT_DATE) + INTERVAL '3 month'
                                            ELSE date_trunc('year', CURRENT_DATE - INTERVAL '1 year') + INTERVAL '3 month'
                                        END,
                                        'DD Mon''YY'
                                    ) || ' to ' ||
                                    TO_CHAR(CURRENT_DATE - INTERVAL '1 day', 'DD Mon''YY') 
                                    AS financial_year_label,

                                    /* ===== period lengths, needed to reproduce query2's
                                       AVG(daily volume)/1411/0.89 -> ROUND to BIGINT step ===== */

                                    (
                                        CASE
                                            WHEN CURRENT_DATE = date_trunc('month', CURRENT_DATE)::DATE
                                            THEN date_trunc('month', CURRENT_DATE)::DATE
                                            ELSE CURRENT_DATE::DATE
                                        END
                                        -
                                        CASE
                                            WHEN CURRENT_DATE = date_trunc('month', CURRENT_DATE)::DATE
                                            THEN date_trunc('month', CURRENT_DATE - INTERVAL '1 month')::DATE
                                            ELSE date_trunc('month', CURRENT_DATE)::DATE
                                        END
                                    ) AS current_period_days,

                                    (
                                        CURRENT_DATE::DATE
                                        -
                                        CASE
                                            WHEN EXTRACT(MONTH FROM CURRENT_DATE) >= 4
                                            THEN (date_trunc('year', CURRENT_DATE) + INTERVAL '3 month')::DATE
                                            ELSE (date_trunc('year', CURRENT_DATE - INTERVAL '1 year') + INTERVAL '3 month')::DATE
                                        END
                                    ) AS ytd_period_days

                                FROM nozzle_sales
                                WHERE zone is not Null AND sap_id LIKE '4%'
                                GROUP BY
                                    zone,
                                    region,
                                    sales_area,
                                    sap_id,
                                    location_name;
                        """
    
    Charts_Connection_Vault_RoutingParams.connection_id = connection_mapping.connection_mapping.get("hpcl_ceg", "1")
    Charts_Connection_Vault_RoutingParams.action = 'execute_query'
    function = await charts_connection_vault_routing(Charts_Connection_Vault_RoutingParams)
    nozzle_sales = await function(query=nozzle_sales_top_query)
    nozzle_sales_df = pl.DataFrame(nozzle_sales)

    # location_details_query = f"""SELECT DISTINCT sap_id, zone, region, sales_area, name FROM location_master where bu = 'RO' """
    # location_data  = await urdhva_base.BasePostgresModel.get_aggr_data(location_details_query, limit=0)
    # loc_df = pl.DataFrame(location_data['data'])

    # nozzle_sales_df = nozzle_sales_df.join(loc_df, on = 'sap_id', how = "left")

    # nozzle_sales_top_df = nozzle_sales_df.filter(
    #     (pl.col("current_ms_kl").is_not_null()) & 
    #     (pl.col("current_ms_kl") != 0)
    # )

    labels = nozzle_sales_df.select([
        "current_date_label",
        "financial_year_label"
    ]).unique().to_dicts()[0]
    print("labels ---->\n", labels)

    current_period_days = nozzle_sales_df.select("current_period_days").unique().to_dicts()[0]["current_period_days"]
    ytd_period_days = nozzle_sales_df.select("ytd_period_days").unique().to_dicts()[0]["ytd_period_days"]

    nozzle_sales_top_df = nozzle_sales_df.filter(pl.col("zone").is_not_null())

    # Outlet-level KL sums (unique per outlet already, no location_master join needed
    # since sap_id is already unique at this grain from the source query).
    nozzle_sales_top_df = (
        nozzle_sales_top_df
        .group_by(["zone", "region", "sales_area", "location_name", "sap_id"])
        .agg([
            pl.sum("current_ms_kl").alias("current_ms_kl"),
            pl.sum("current_power_kl").alias("current_power_kl"),
            pl.sum("ytd_power_kl").alias("ytd_power_kl"),
            pl.sum("ytd_ms_kl").alias("ytd_ms_kl"),
        ])
        .with_columns([
            (pl.col("current_ms_kl") + pl.col("current_power_kl")).alias("current_ms_total_kl"),
            (pl.col("ytd_power_kl") + pl.col("ytd_ms_kl")).alias("ytd_ms_total_kl"),
        ])
    )

    def with_unit_conversion_pct(df: pl.DataFrame, normal_col: str, branded_col: str,
                                  days: int, prefix: str) -> pl.DataFrame:
        """
        Reproduces query2's methodology exactly:
          units = ROUND( (period_KL_sum * 1000 / days) / 1411.0 / 0.89 )   -- round-half-up, like Postgres ROUND()::BIGINT
          pct   = branded_units * 100 / (normal_units + branded_units)
        This must be applied AFTER summing KL at whatever level (outlet / sales_area /
        region / zone) -- not before -- since the rounding to whole units is what
        query2 does, and it's exactly what makes 12.7322% collapse to 12.71%.
        """
        return df.with_columns([
            ((pl.col(normal_col).cast(pl.Float64) * 1000.0 / days / 1411.0 / 0.89) + 0.5)
                .floor().alias(f"{prefix}_normal_units"),
            ((pl.col(branded_col).cast(pl.Float64) * 1000.0 / days / 1411.0 / 0.89) + 0.5)
                .floor().alias(f"{prefix}_branded_units"),
        ]).with_columns([
            (
                pl.col(f"{prefix}_branded_units") * 100.0
                / (pl.col(f"{prefix}_normal_units") + pl.col(f"{prefix}_branded_units"))
            ).round(2).alias(f"{prefix}_conversion_pct")
        ])

    # ---- Outlet level ----
    # NOTE: the units-rounding formula (with_unit_conversion_pct) was reverse-engineered
    # from query2, which only ever aggregates at zone / Pan-India level. At individual
    # outlet volumes, dividing by 1411/0.89 and rounding to a whole number frequently
    # collapses BOTH normal_units and branded_units to 0, giving 0/0 = nan. There's no
    # outlet-level reference to validate the units approach against anyway (the
    # dashboard only shows zone/region/sales_area tables), so outlet-level %Conversion
    # is computed as a plain ratio of the KL sums instead.
    nozzle_sales_top_df = nozzle_sales_top_df.with_columns([
        (pl.col("current_power_kl").cast(pl.Float64) / pl.col("current_ms_total_kl").cast(pl.Float64) * 100)
            .round(2).alias("current_conversion_pct"),
        (pl.col("ytd_power_kl").cast(pl.Float64) / pl.col("ytd_ms_total_kl").cast(pl.Float64) * 100)
            .round(2).alias("ytd_conversion_pct"),
    ])

    top_3_retail_outlets = (
        nozzle_sales_top_df
        .filter(pl.col("current_conversion_pct").is_not_null() & pl.col("current_conversion_pct").is_not_nan())
        .sort("current_conversion_pct", descending=True)
        .head(3)
    )

    # ---- Sales area level ----
    top_3_sales_areas = (
        nozzle_sales_top_df
        .group_by(["zone", "region", "sales_area"])
        .agg([
            pl.sum("current_ms_kl").alias("current_ms_kl"),
            pl.sum("current_power_kl").alias("current_power_kl_sum"),
            pl.sum("ytd_ms_kl").alias("ytd_ms_kl"),
            pl.sum("ytd_power_kl").alias("ytd_power_kl_sum"),
        ]).with_columns([
            (pl.col("current_ms_kl") + pl.col("current_power_kl_sum")).alias("current_ms_total_kl_sum"),
            (pl.col("ytd_ms_kl") + pl.col("ytd_power_kl_sum")).alias("ytd_ms_total_kl_sum"),
        ])
    )
    top_3_sales_areas = with_unit_conversion_pct(
        top_3_sales_areas, "current_ms_kl", "current_power_kl_sum", current_period_days, "current"
    )
    top_3_sales_areas = with_unit_conversion_pct(
        top_3_sales_areas, "ytd_ms_kl", "ytd_power_kl_sum", ytd_period_days, "ytd"
    )
    top_3_sales_areas = (
        top_3_sales_areas
        .filter(pl.col("current_conversion_pct").is_not_null() & pl.col("current_conversion_pct").is_not_nan())
        .sort("current_conversion_pct", descending=True)
        .head(3)
    )

    # ---- Region level ----
    top_3_regions = (
        nozzle_sales_top_df
        .group_by(["zone", "region"])
        .agg([
            pl.sum("current_ms_kl").alias("current_ms_kl"),
            pl.sum("current_power_kl").alias("current_power_kl_sum"),
            pl.sum("ytd_ms_kl").alias("ytd_ms_kl"),
            pl.sum("ytd_power_kl").alias("ytd_power_kl_sum"),
        ]).with_columns([
            (pl.col("current_ms_kl") + pl.col("current_power_kl_sum")).alias("current_ms_total_kl_sum"),
            (pl.col("ytd_ms_kl") + pl.col("ytd_power_kl_sum")).alias("ytd_ms_total_kl_sum"),
        ])
    )
    top_3_regions = with_unit_conversion_pct(
        top_3_regions, "current_ms_kl", "current_power_kl_sum", current_period_days, "current"
    )
    top_3_regions = with_unit_conversion_pct(
        top_3_regions, "ytd_ms_kl", "ytd_power_kl_sum", ytd_period_days, "ytd"
    )
    top_3_regions = (
        top_3_regions
        .filter(pl.col("current_conversion_pct").is_not_null() & pl.col("current_conversion_pct").is_not_nan())
        .sort("current_conversion_pct", descending=True)
        .head(3)
    )

    # ---- Zone level ----
    top_3_zones = (
        nozzle_sales_top_df
        .filter(pl.col("zone").is_not_null())
        .group_by(["zone"])
        .agg([
            pl.sum("current_ms_kl").alias("current_ms_kl"),
            pl.sum("current_power_kl").alias("current_power_kl_sum"),
            pl.sum("ytd_ms_kl").alias("ytd_ms_kl"),
            pl.sum("ytd_power_kl").alias("ytd_power_kl_sum"),
        ]).with_columns([
            (pl.col("current_ms_kl") + pl.col("current_power_kl_sum")).alias("current_ms_total_kl_sum"),
            (pl.col("ytd_ms_kl") + pl.col("ytd_power_kl_sum")).alias("ytd_ms_total_kl_sum"),
        ])
    )
    top_3_zones = with_unit_conversion_pct(
        top_3_zones, "current_ms_kl", "current_power_kl_sum", current_period_days, "current"
    )
    top_3_zones = with_unit_conversion_pct(
        top_3_zones, "ytd_ms_kl", "ytd_power_kl_sum", ytd_period_days, "ytd"
    )
    top_3_zones = (
        top_3_zones
        .filter(pl.col("current_conversion_pct").is_not_null() & pl.col("current_conversion_pct").is_not_nan())
        .sort("current_conversion_pct", descending=True)
        .head(3)
    )
    return {
        "top_3_retail_outlets": top_3_retail_outlets.to_dicts(),
        "top_3_sales_areas": top_3_sales_areas.to_dicts(),
        "top_3_regions": top_3_regions.to_dicts(),
        "top_3_zones": top_3_zones.to_dicts(),
        "nozzle_present_month": labels["current_date_label"],
        "nozzle_previous_month": labels["financial_year_label"]
    }
